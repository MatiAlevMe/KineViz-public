import tempfile
import shutil # Añadir para borrar directorio
import logging
from pathlib import Path
from datetime import datetime
import numpy as np
import pandas as pd

from typing import Optional # Importar Optional

# Importar servicios y helpers
from .file_service import FileService
from .study_service import StudyService
from kineviz.ui.widgets import charting  # Importar nuestro módulo de gráficos
from kineviz.core.data_processing import processors # Importar processors
# Importar el validador de nombres de archivo
from kineviz.ui.utils.validators import validate_filename_for_study_criteria

from kineviz.config.settings import AppSettings # Import AppSettings
from kineviz.core.undo_manager import UndoManager # Import UndoManager

# Importar reportlab
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image,
                                Table, TableStyle)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
# import itertools # No usado directamente aquí
import json
import time # Añadir import


# Define logger before it's potentially used in the except block below
logger = logging.getLogger(__name__)

from kineviz.core.backup_manager import create_backup # Import for automatic backups

# Importar scipy para tests estadísticos
try:
    from scipy import stats
except ImportError:
    logger.warning("Scipy no está instalado. Las pruebas estadísticas discretas no estarán disponibles.")
    stats = None

# Importar spm1d para análisis continuo
try:
    import spm1d
except ImportError:
    logger.warning("spm1d no está instalado. El análisis continuo no estará disponible.")
    spm1d = None


# Importar openpyxl para Excel (opcional)
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.info("openpyxl no está instalado. La exportación a .xlsx no estará disponible.")


class AnalysisService:
    # Eliminar app_settings de la inicialización
    def __init__(self, study_service: StudyService, file_service: FileService, settings: AppSettings, undo_manager: UndoManager):
        """
        Inicializa el AnalysisService.

        :param study_service: Instancia de StudyService.
        :param file_service: Instancia de FileService.
        :param settings: Instancia de AppSettings.
        :param undo_manager: Instancia de UndoManager.
        """
        self.study_service = study_service
        self.file_service = file_service
        self.settings = settings # Use passed AppSettings instance
        self.undo_manager = undo_manager # Use passed UndoManager instance

    def get_analysis_parameters(self, study_id: int) -> dict:
        """
        Obtiene los parámetros disponibles para análisis de un estudio, incluyendo cálculos.

        :param study_id: ID del estudio.
        :return: Diccionario con sets de parámetros disponibles
                 {'patients': set(), 'data_types': set(), 'sub_values_by_vi': dict, 'calculations': set()}
                 Retorna sets vacíos si no se encuentran parámetros o hay error.
        """
        try:
            # Obtener parámetros únicos del FileService (ahora devuelve 'sub_values_by_vi')
            params = self.file_service.get_unique_study_parameters(study_id)
            # Añadir cálculos fijos
            params['calculations'] = {'Maximo', 'Minimo', 'Rango'}
            # Asegurar que 'sub_values_by_vi' exista aunque esté vacío
            if 'sub_values_by_vi' not in params:
                params['sub_values_by_vi'] = {}
            # Renombrar 'frequencies' a 'data_types' si existe para consistencia interna de este método
            if 'frequencies' in params:
                params['data_types'] = params.pop('frequencies')
            elif 'data_types' not in params: # Asegurar que exista
                 params['data_types'] = set()
            return params
        except Exception as e:
            logger.error(f"Error obteniendo parámetros de análisis para estudio {study_id}: {e}", exc_info=True)
            # Devolver vacío en caso de error para que la UI no falle
            return {'patients': set(), 'data_types': set(), # Asegurar data_types aquí también
                    'sub_values_by_vi': dict(), 'calculations': set()}

    def get_available_frequencies_for_study(self, study_id: int) -> list[str]:
        """
        Obtiene los tipos de dato disponibles. Para el análisis continuo,
        actualmente solo se soporta y devuelve "Cinematica" si está presente.

        :param study_id: ID del estudio.
        :return: Lista con ["Cinematica"] si datos cinemáticos existen, sino lista vacía.
        """
        try:
            params = self.file_service.get_unique_study_parameters(study_id)
            # 'frequencies' es la clave que get_unique_study_parameters usa actualmente
            available_data_types = params.get('frequencies', set())
            if "Cinematica" in available_data_types:
                return ["Cinematica"]
            return []
        except Exception as e:
            logger.error(f"Error obteniendo tipos de dato disponibles para estudio {study_id}: {e}", exc_info=True)
            return []

    def get_data_columns_for_frequency(self, study_id: int, frequency: str) -> list[str]:
        """
        Obtiene las columnas de datos (variables) disponibles para un Tipo de Dato específico
        de un estudio, en formato "Atributo/Columna/Unidad".
        Excluye columnas comunes no analizables como 'Frame', 'Sub Frame', 'Tiempo'.
        El parámetro 'frequency' aquí se interpreta como 'Tipo de Dato'.

        :param study_id: ID del estudio.
        :param frequency: Tipo de Dato seleccionada (ej: "Cinematica").
        :return: Lista ordenada de nombres de columnas de datos formateadas.
                 Retorna lista vacía si no hay archivos, columnas o hay error.
        """
        logger.debug(f"Obteniendo columnas de datos para estudio {study_id}, Tipo de Dato '{frequency}'")
        try:
            processed_files, _ = self.file_service.get_study_files(
                study_id=study_id, page=1, per_page=1,
                file_type='Processed', frequency=frequency # FileService espera 'frequency'
            )

            if not processed_files:
                logger.warning(f"No se encontraron archivos procesados para Tipo de Dato '{frequency}' en estudio {study_id}.")
                return []

            sample_file_path = processed_files[0]['path']
            logger.debug(f"Leyendo cabeceras del archivo de muestra: {sample_file_path}")

            if not sample_file_path.exists():
                logger.error(f"El archivo de muestra {sample_file_path} no existe.")
                return []

            headers = self._parse_processed_file_headers(sample_file_path)
            if not headers:
                logger.warning(f"No se pudieron parsear las cabeceras de {sample_file_path.name} para el Tipo de Dato '{frequency}'.")
                return []

            atributos, columnas, unidades = headers

            formatted_columns = []
            # Excluir columnas no deseadas (insensible a mayúsculas/minúsculas)
            excluded_cols_lower = {"frame", "sub frame", "tiempo"}
            
            # Mantener un registro del último atributo válido para propagar en caso de celdas combinadas
            last_valid_attribute = "N/A"

            for i in range(len(columnas)): # Iterar sobre el número de columnas de datos reales
                col_name_cleaned = columnas[i].strip()
                col_name_lower = col_name_cleaned.lower()

                if col_name_lower not in excluded_cols_lower and col_name_cleaned:
                    # Determinar el atributo a usar
                    current_attribute = atributos[i].strip()
                    if current_attribute:
                        last_valid_attribute = current_attribute
                    # Si el atributo actual está vacío pero la columna no, usar el último válido
                    elif col_name_cleaned: # Solo usar last_valid_attribute si hay un nombre de columna
                        current_attribute = last_valid_attribute
                    else: # Si tanto atributo como columna están vacíos, es raro, usar N/A
                        current_attribute = "N/A"
                        
                    unit = unidades[i].strip() if unidades[i].strip() else "s.u." # sin unidad
                    
                    formatted_columns.append(f"{current_attribute}/{col_name_cleaned}/{unit}")
            
            logger.info(f"Columnas de datos formateadas encontradas para Tipo de Dato '{frequency}' (después de filtro): {formatted_columns}")
            return sorted(list(set(formatted_columns))) # Usar set para asegurar unicidad y luego ordenar

        except Exception as e:
            logger.error(f"Error obteniendo columnas de datos para estudio {study_id}, Tipo de Dato '{frequency}': {e}", exc_info=True)
            return []

    def _read_processed_file_data(self, file_path: Path) -> pd.DataFrame | None:
        """
        Lee los datos numéricos de un archivo procesado (.txt separado por ';').
        Omite las primeras 4 líneas de encabezado y las últimas 3 de estadísticas.
        Devuelve un DataFrame con los datos numéricos o None si hay error.
        """
        try:
            # Leer todas las líneas primero para poder omitir las últimas
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            # 4 header + 3 stats = 7. Necesita al menos 1 fila de datos.
            if len(lines) <= 7:
                logger.warning(f"Archivo {file_path.name} no contiene "
                               f"suficientes líneas para extraer datos.")
                return None

            # Omitir encabezado y estadísticas
            data_lines = lines[4:-3]

            # Usar pandas para leer los datos, especificando el separador
            # Necesitamos pasar las líneas como un stream
            from io import StringIO
            data_io = StringIO("".join(data_lines))

            # --- Determinar número de columnas de los datos reales ---
            if not data_lines:
                logger.warning(f"Archivo {file_path.name} no contiene líneas "
                               f"de datos después de quitar cabecera/pie.")
                return None
            first_data_line_parts = data_lines[0].strip().split(';')
            num_data_cols = len(first_data_line_parts)
            # Add debug for first data line content
            logger.debug(f"Primera línea de datos: '{data_lines[0].strip()}'")
            logger.debug(f"Detectadas {num_data_cols} columnas en la primera "
                         f"línea de datos de {file_path.name}")

            if num_data_cols == 0:
                logger.warning(f"No se detectaron columnas de datos en "
                               f"{file_path.name}")
                return None

            # --- Generar nombres de columna basados en línea 3 (';' sep) ---
            # Leer la línea de nombres de columna (índice 2)
            col_names_line = lines[2].strip()
            raw_col_names_from_header = col_names_line.split(';')
            logger.debug(f"Nombres crudos leídos de línea 3: {raw_col_names_from_header}")

            # Validar si el número de nombres coincide con datos
            if len(raw_col_names_from_header) != num_data_cols:
                logger.warning(f"Discrepancia en {file_path.name}: "
                               f"Nombres línea 3 ({len(raw_col_names_from_header)}) != "
                               f"Columnas datos ({num_data_cols}). "
                               f"Se usarán nombres truncados/rellenados.")
                # Podríamos fallar aquí si es grave

            # Sanear nombres de columna leídos (unicidad, no vacíos)
            final_col_names = []
            counts = {}
            # Sanitize names from the header line
            for i, name_from_header_line3 in enumerate(raw_col_names_from_header):
                current_col_name_for_df = name_from_header_line3.strip() # Start with stripped name

                # Apply prefix cleaning only to actual data columns (index >= 3)
                # This ensures consistency with _parse_processed_file_headers
                if i >= 3: # Indices 0, 1, 2 are Frame, Sub Frame, Tiempo
                    if ':' in current_col_name_for_df:
                        name_parts = current_col_name_for_df.split(':', 1)
                        if len(name_parts) > 1:
                            current_col_name_for_df = name_parts[1].strip()
                        # else: keep original stripped name if split fails unexpectedly (e.g. "Var:X" but no prefix part)
                
                # Handle empty names (after potential prefix stripping)
                if not current_col_name_for_df:
                    current_col_name_for_df = f"Unnamed_{i}"
                
                # Make unique if necessary
                temp_unique_name = current_col_name_for_df
                if temp_unique_name in counts:
                    counts[temp_unique_name] += 1
                    unique_name_for_df = f"{temp_unique_name}_{counts[temp_unique_name]}"
                else:
                    counts[temp_unique_name] = 0
                    unique_name_for_df = temp_unique_name
                final_col_names.append(unique_name_for_df)
            
            logger.debug(f"Nombres de columna saneados para DataFrame ({len(final_col_names)}): {final_col_names}")

            # Ajustar lista final si hubo discrepancia con num_data_cols
            if len(final_col_names) > num_data_cols:
                final_col_names = final_col_names[:num_data_cols]  # Truncar
            elif len(final_col_names) < num_data_cols:
                # Pad con nombres genéricos
                for i in range(len(final_col_names), num_data_cols):
                    final_col_names.append(f"Data_Col_{i}")

            logger.debug(f"Nombres columna finales ajustados para "
                         f"{file_path.name} ({len(final_col_names)}): "
                         f"{final_col_names}")
            # --- Fin ajuste de nombres ---

            try:
                df = pd.read_csv(data_io, sep=';', header=None,
                                 names=final_col_names, na_values=[''],
                                 keep_default_na=True)
            except pd.errors.ParserError as pe:
                # Añadir más contexto al error de pandas
                logger.error(f"Error Pandas al parsear {file_path.name} con "
                             f"{len(final_col_names)} columnas esperadas: {pe}",
                             exc_info=True)
                raise  # Relanzar para que se maneje en el bloque exterior

            # Seleccionar solo columnas numéricas
            numeric_cols = []
            for col in df.columns:
                # Intentar convertir a numérico, ignorando 'Tiempo'
                if col.lower() == 'tiempo':
                    numeric_cols.append(col)
                    continue
                try:
                    pd.to_numeric(df[col])
                    numeric_cols.append(col)
                except (ValueError, TypeError):
                    pass  # Ignorar columnas no numéricas

            return df[numeric_cols]

        except FileNotFoundError:
            logger.error(f"Archivo no encontrado al leer datos: {file_path}")
            return None
        except Exception as e:
            logger.error(f"Error leyendo datos de {file_path.name}: {e}",
                         exc_info=True)
            return None

    def _get_data_for_parameters(self, study_id: int, parameters: dict) -> dict:
        """
        Obtiene y estructura los datos numéricos de los archivos que coinciden
        con los parámetros seleccionados.

        :param study_id: ID del estudio.
        :param parameters: Diccionario con listas de 'patients', 'frequencies', 'descriptors', 'calculations'.
        :return: Diccionario anidado:
                 {
                     'frequency1': {
                         'descriptor_combo_key': { # Clave basada en sub-valores encontrados
                             'patient1': DataFrame,
                             'patient2': DataFrame, ...
                         }, ...
                     }, ...
                 Retorna diccionario vacío si no hay datos o hay error.
        """
        logger.debug(f"Inicio _get_data_for_parameters para estudio {study_id} con params: {parameters}") # LOG INICIO
        structured_data = {}
        # Usar método protegido para obtener ruta
        study_path = self.file_service._get_study_path(study_id)
        if not study_path:
            return {}

        selected_patients = parameters.get('patients', [])
        selected_frequencies = parameters.get('frequencies', [])
        # selected_descriptors ya no se usa directamente aquí, se usa la estructura VI

        # Obtener estructura de VIs del estudio para validación
        try:
            study_details = self.study_service.get_study_details(study_id)
            independent_variables = study_details.get('independent_variables', [])
        except Exception as e:
            logger.error(f"Error obteniendo VIs estudio {study_id}: {e}",
                         exc_info=True)
            return {}

        for patient in selected_patients:
            patient_path = study_path / patient
            if not patient_path.is_dir():
                continue

            for freq in selected_frequencies:
                freq_path = patient_path / freq
                if not freq_path.is_dir():
                    continue

                if freq not in structured_data:
                    structured_data[freq] = {}

                # Iterar sobre archivos en carpeta de frecuencia
                # Asumiendo extensión .txt para procesados
                for file_path in freq_path.glob('*.txt'):
                    filename = file_path.name

                    # Validar nombre de archivo usando VIs y desempaquetar 4 valores
                    is_valid_name, _, extracted_descriptors, _ = validate_filename_for_study_criteria(
                        filename, independent_variables
                    )
                    # Solo necesitamos is_valid_name y extracted_descriptors aquí
                    if not is_valid_name:
                        continue # Omitir archivo si no cumple criterios

                    # Crear clave de grupo combinada basada en VIs y sub-valores extraídos
                    # Formato: "VI1=DescA;VI2=DescB" o "VI1=Nulo;VI2=DescC"
                    group_parts = []
                    for i, desc in enumerate(extracted_descriptors):
                        vi_name = independent_variables[i].get('name', f'VI{i+1}') # Usar nombre VI
                        value = desc if desc is not None else "Nulo" # Usar "Nulo" si es None
                        group_parts.append(f"{vi_name}={value}")

                    # Usar ';' como separador para evitar conflictos con nombres
                    group_key = ";".join(group_parts) if group_parts else "SinGrupo"

                    if group_key not in structured_data[freq]:
                        structured_data[freq][group_key] = {}

                    # Leer datos del archivo
                    # (La lógica de lectura y concatenación permanece igual)
                    df_data = self._read_processed_file_data(file_path)
                    if df_data is not None and not df_data.empty:
                        # Acumular datos si ya existe una entrada para este paciente/freq/group_key
                        if patient not in structured_data[freq][group_key]:
                            structured_data[freq][group_key][patient] = df_data
                        else:
                            # Concatenar DataFrames
                            structured_data[freq][group_key][patient] = \
                                pd.concat(
                                    [structured_data[freq][group_key][patient], df_data],
                                    ignore_index=True
                                )
                    else:
                        logger.warning(f"No se pudieron leer datos válidos de {filename}")

        return structured_data

    def _calculate_statistic(self, df: pd.DataFrame, calculation: str) -> pd.Series | None:
        """Calcula una estadística ('Maximo', 'Minimo', 'Rango') para cada columna numérica del DataFrame."""
        if df is None or df.empty:
            return None

        # Seleccionar solo columnas numéricas (excluir 'Tiempo')
        numeric_df = df.select_dtypes(include=np.number)
        if 'Tiempo' in numeric_df.columns:
            numeric_df = numeric_df.drop(columns=['Tiempo'])

        # Devolver None si no quedan columnas numéricas O si todas son NaN
        if numeric_df.empty or numeric_df.isnull().all().all():
            return None

        if calculation == "Maximo":
            return numeric_df.max(skipna=True)
        elif calculation == "Minimo":
            return numeric_df.min(skipna=True)
        elif calculation == "Rango":
            return numeric_df.max(skipna=True) - numeric_df.min(skipna=True)
        else:
            logger.warning(f"Cálculo no soportado '{calculation}'")
            return None

    def perform_analysis(self, study_id: int, parameters: dict):
        """
        Realiza un análisis basado en los parámetros proporcionados.
        Calcula las estadísticas seleccionadas para los datos agrupados.

        :param study_id: ID del estudio a analizar.
        :param parameters: Diccionario con los parámetros de análisis ('patients', 'frequencies', 'descriptors', 'calculations').
                 frecuencia -> descriptor_key -> calculo -> paciente -> Serie.
                 Ej: {'Cinematica': {'CMJ_PRE': {'Maximo': {'P01': pd.Series}}}}
        """
        logger.info(f"Realizando análisis para estudio {study_id} con "
                    f"parámetros: {parameters}")
        structured_data = self._get_data_for_parameters(study_id, parameters)
        analysis_results = {}
        selected_calculations = parameters.get('calculations', [])

        if not structured_data:
            logger.warning(f"No se encontraron datos para los parámetros "
                           f"seleccionados en estudio {study_id}.")
            return {}

        for freq, group_data in structured_data.items(): # Cambiar descriptor_data a group_data
            analysis_results[freq] = {}
            for group_key, patient_data in group_data.items(): # Cambiar descriptor_key a group_key
                analysis_results[freq][group_key] = {}
                for calc in selected_calculations:
                    analysis_results[freq][group_key][calc] = {}
                    for patient, df in patient_data.items():
                        stats = self._calculate_statistic(df, calc)
                        if stats is not None:
                            analysis_results[freq][group_key][calc][patient] = stats

        logger.info(f"Análisis completado para estudio {study_id}.")
        return analysis_results


    def generate_report(self, study_id: int, parameters: dict, output_path_str: str):
        """
        Genera un reporte PDF del análisis.

        :param study_id: ID del estudio.
        :param parameters: Parámetros del análisis.
        :param output_path_str: Ruta (string) donde guardar el reporte PDF.
        """
        logger.info(f"Generando reporte para estudio {study_id} en "
                    f"{output_path_str}...")
        output_path = Path(output_path_str)
        # Asegurar que directorio exista
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # --- Obtener Datos y Detalles ---
        try:
            study_details = self.study_service.get_study_details(study_id)
            study_name = study_details.get('name', f'Estudio {study_id}')
        except Exception as e:
            raise ValueError(f"No se pudieron obtener detalles del estudio "
                             f"{study_id}: {e}")

        # Obtener datos estructurados (agrupados por paciente)
        structured_data = self._get_data_for_parameters(study_id, parameters)
        selected_calculations = parameters.get('calculations', [])
        selected_patients = parameters.get('patients', [])

        if not structured_data:
            raise ValueError("No se encontraron datos para generar el reporte "
                             "con los parámetros seleccionados.")

        # --- Crear Directorio Temporal para Gráficos ---
        # Usar tempfile para limpieza automática si falla
        with tempfile.TemporaryDirectory(
                prefix=f"kineviz_report_{study_id}_") as temp_dir_str:
            temp_dir = Path(temp_dir_str)
            logger.debug(f"Directorio temporal para gráficos: {temp_dir}")

            # --- Configurar PDF con ReportLab ---
            doc = SimpleDocTemplate(
                output_path_str, pagesize=letter,
                leftMargin=0.75*inch, rightMargin=0.75*inch,
                topMargin=1*inch, bottomMargin=1*inch
            )
            styles = getSampleStyleSheet()
            story = []

            # --- Título y Metadatos ---
            story.append(Paragraph(f"Reporte de Análisis - {study_name}",
                                   styles['h1']))
            story.append(Spacer(1, 0.2*inch))
            story.append(Paragraph(f"Fecha de Generación: "
                                   f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                                   styles['Normal']))
            story.append(Spacer(1, 0.1*inch))

            # Parámetros Usados
            story.append(Paragraph("<b>Parámetros Seleccionados:</b>",
                                   styles['h3']))
            # Obtener alias del estudio
            study_aliases = self.study_service.get_study_aliases(study_id)
            # Mostrar parámetros seleccionados (ya no incluye 'descriptors' directamente)
            param_text = (
                f"<b>Participantes:</b> {', '.join(parameters.get('patients', []))}<br/>"
                f"<b>Tipos de Datos:</b> {', '.join(parameters.get('frequencies', []))}<br/>"
                # Podríamos añadir VIs/Sub-valores si se seleccionaron explícitamente,
                # pero por ahora omitimos esa parte ya que el análisis agrupa por todas las combinaciones.
                f"<b>Cálculos:</b> {', '.join(parameters.get('calculations', []))}"
            )
            story.append(Paragraph(param_text, styles['BodyText']))
            story.append(Spacer(1, 0.3*inch))

            # --- Iterar y Generar Contenido ---
            plot_counter = 0
            for freq, descriptor_data in structured_data.items():
                story.append(Paragraph(f"Resultados para Tipos de Dato: {freq}",
                                       styles['h2']))
                story.append(Spacer(1, 0.1*inch))

                for group_key, patient_data in descriptor_data.items(): # Usar descriptor_data
                    # Convertir group_key a formato legible con alias
                    group_display_parts = []
                    if group_key != "SinGrupo":
                        for part in group_key.split(';'):
                            vi_name, desc_value = part.split('=', 1)
                            alias = study_aliases.get(desc_value, desc_value) # Aplicar alias
                            group_display_parts.append(f"{vi_name}: {alias}")
                    group_display = ", ".join(group_display_parts) if group_display_parts else "Grupo General"

                    story.append(Paragraph(f"Grupo: {group_display}", styles['h3']))
                    story.append(Spacer(1, 0.1*inch))

                    # --- Boxplot General por Paciente ---
                    boxplot_data = {}
                    all_numeric_columns = set()
                    for patient, df in patient_data.items():
                        numeric_df = df.select_dtypes(include=np.number)
                        if 'Tiempo' in numeric_df.columns:
                            numeric_df = numeric_df.drop(columns=['Tiempo'])
                        if not numeric_df.empty:
                            # Usar todos los valores de todas las columnas
                            boxplot_data[patient] = numeric_df.values.flatten()
                            all_numeric_columns.update(numeric_df.columns)

                    if boxplot_data:
                        plot_counter += 1
                        boxplot_filename = temp_dir / f"boxplot_{plot_counter}.png"
                        # Usar group_display (con alias) en el título del gráfico
                        charting.create_boxplot(
                            data_dict=boxplot_data,
                            title=f"Distribución General - {freq} ({group_display})",
                            ylabel="Valor Medición",
                            output_path=boxplot_filename
                        )
                        if boxplot_filename.exists():
                            # Ajustar tamaño
                            story.append(Image(str(boxplot_filename),
                                               width=6*inch, height=4*inch))
                            story.append(Spacer(1, 0.2*inch))
                        else:
                            story.append(Paragraph(f"<i>Error generando boxplot "
                                                   f"{plot_counter}</i>",
                                                   styles['Italic']))
                    else:
                        story.append(Paragraph("<i>No hay datos suficientes "
                                               "para el boxplot general.</i>",
                                               styles['Italic']))

                    # --- Cálculos y Gráficos de Barras ---
                    for calc in selected_calculations:
                        story.append(Paragraph(f"<b>Cálculo: {calc}</b>",
                                               styles['Normal']))
                        story.append(Spacer(1, 0.05*inch))

                        calc_results_by_patient = {}
                        valid_columns_for_calc = set()
                        for patient, df in patient_data.items():
                            stats = self._calculate_statistic(df, calc)
                            if stats is not None and not stats.empty:
                                calc_results_by_patient[patient] = stats
                                valid_columns_for_calc.update(stats.index)

                        if calc_results_by_patient:
                            # --- Gráfico de Barras (Promedio por Paciente) ---
                            # Calcular promedio del cálculo para cada paciente
                            avg_calc_per_patient = {
                                patient: results.mean(skipna=True)
                                for patient, results
                                in calc_results_by_patient.items()
                                if results is not None
                            }
                            # Filtrar pacientes sin promedio válido
                            valid_avg_calc = {p: v for p, v
                                              in avg_calc_per_patient.items()
                                              if pd.notna(v)}

                            if valid_avg_calc:
                                plot_counter += 1
                                barchart_filename = temp_dir / f"barchart_{plot_counter}.png"
                                # Usar group_display (con alias) en el título del gráfico
                                charting.create_barchart(
                                    data_dict=valid_avg_calc,
                                    title=f"{calc} Promedio - {freq} ({group_display})",
                                    xlabel="Participante",
                                    ylabel=f"{calc} Promedio",
                                    output_path=barchart_filename
                                )
                                if barchart_filename.exists():
                                    story.append(Image(str(barchart_filename),
                                                       width=6*inch, height=4*inch))
                                    story.append(Spacer(1, 0.1*inch))
                                else:
                                    story.append(Paragraph(f"<i>Error generando "
                                                           f"barchart {plot_counter}</i>",
                                                           styles['Italic']))
                            else:
                                story.append(Paragraph(f"<i>No hay datos suficientes "
                                                       f"para el gráfico de barras "
                                                       f"de {calc}.</i>",
                                                       styles['Italic']))

                            # --- Tabla de Resultados Detallados ---
                            # Crear tabla: pacientes (filas) x mediciones (cols)
                            if valid_columns_for_calc:
                                sorted_columns = sorted(list(valid_columns_for_calc))
                                table_data = [['Participante'] + sorted_columns]  # Cabecera
                                # Iterar en orden de selección
                                for patient in selected_patients:
                                    if patient in calc_results_by_patient:
                                        results = calc_results_by_patient[patient]
                                        row_values = [
                                            f"{results.get(col, 'N/A'):.3f}"
                                            if pd.notna(results.get(col)) else 'N/A'
                                            for col in sorted_columns
                                        ]
                                        row = [patient] + row_values
                                        table_data.append(row)
                                    # else: # Opcional: fila si no hay datos
                                    #     table_data.append([patient] + ['N/A'] * len(sorted_columns))

                                # Si hay datos además de la cabecera
                                if len(table_data) > 1:
                                    table = Table(table_data, hAlign='LEFT')
                                    table.setStyle(TableStyle([
                                        ('BACKGROUND', (0, 0), (-1, 0),
                                         colors.grey),
                                        ('TEXTCOLOR', (0, 0), (-1, 0),
                                         colors.whitesmoke),
                                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                        ('FONTNAME', (0, 0), (-1, 0),
                                         'Helvetica-Bold'),
                                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                                        ('BACKGROUND', (0, 1), (-1, -1),
                                         colors.beige),
                                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                                    ]))
                                    story.append(table)
                                    story.append(Spacer(1, 0.2*inch))
                                else:
                                    story.append(Paragraph(f"<i>No hay resultados "
                                                           f"detallados para {calc}.</i>",
                                                           styles['Italic']))
                            else:
                                story.append(Paragraph(f"<i>No hay columnas válidas "
                                                       f"para mostrar resultados "
                                                       f"detallados de {calc}.</i>",
                                                       styles['Italic']))
                        else:
                            story.append(Paragraph(f"<i>No se pudieron calcular "
                                                   f"resultados para {calc}.</i>",
                                                   styles['Italic']))
                        # Espacio después de cada cálculo
                        story.append(Spacer(1, 0.1*inch))

                    # Espacio después de cada tipo/periodo
                    story.append(Spacer(1, 0.2*inch))

            # --- Construir PDF ---
            try:
                doc.build(story)
                logger.info(f"Reporte PDF generado exitosamente en {output_path}")
            except Exception as build_e:
                logger.error(f"Error construyendo el PDF para estudio "
                             f"{study_id}: {build_e}", exc_info=True)
                raise  # Relanzar error de construcción

        # El directorio temporal se limpia automáticamente al salir del 'with'

    # --- Métodos para Gestión de Reportes ---

    def list_reports(self, study_id: int) -> list[dict]:
        """
        Lista los archivos PDF de reportes generados para un estudio.

        :param study_id: ID del estudio.
        :return: Lista de diccionarios {'name': str, 'path': str}.
                 Retorna lista vacía si no hay reportes o hay error.
        """
        reports = []
        try:
            study_path = self.file_service._get_study_path(study_id)
            if not study_path:
                return []

            reports_dir = study_path / "reportes"
            if reports_dir.exists() and reports_dir.is_dir():
                for item in reports_dir.glob("*.pdf"):
                    if item.is_file():
                        reports.append({'name': item.name, 'path': str(item)})
                # Ordenar por nombre (o fecha si se extrae)
                reports.sort(key=lambda x: x['name'], reverse=True)
        except Exception as e:
            logger.error(f"Error listando reportes para estudio {study_id}: {e}",
                         exc_info=True)
        return reports

    def delete_report(self, report_path_str: str):
        """
        Elimina un archivo de reporte específico.

        :param report_path_str: Ruta completa (string) del archivo PDF a eliminar.
        :raises FileNotFoundError: Si el archivo no existe.
        :raises OSError: Si ocurre un error al eliminar.
        """
        report_path = Path(report_path_str)
        if not report_path.exists():
            raise FileNotFoundError(f"El archivo de reporte no existe: "
                                    f"{report_path}")
        if not report_path.is_file():
            raise ValueError(f"La ruta no es un archivo: {report_path}")

        if self.undo_manager.is_undo_enabled():
            if not self.undo_manager.prepare_undo_cache():
                logger.error(f"Failed to prepare undo cache for deleting report {report_path}. Aborting delete operation.")
                raise Exception(f"Failed to prepare undo cache for deleting report {report_path}. Deletion aborted.")
            
            if not self.undo_manager.cache_item_for_undo(str(report_path), "analysis_report_file"):
                logger.warning(f"Failed to cache report file {report_path} for undo. Deletion will proceed but undo might be partial.")

        try:
            report_path.unlink()
            logger.info(f"Reporte eliminado: {report_path}")
        except OSError as e:
            logger.error(f"Error al eliminar el reporte {report_path}: {e}",
                         exc_info=True)
            # If unlink fails, the undo cache might be prepared. It will be cleared on the next prepare.
            raise

    # --- Métodos para Análisis Continuo (Fase 5) ---

    def _get_normalized_data_for_groups(self, study_id: int, config: dict) -> dict[str, list[np.ndarray]]:
        """
        Prepara los datos normalizados para el análisis continuo.

        :param study_id: ID del estudio.
        :param config: Diccionario de configuración del ContinuousAnalysisConfigDialog.
                       Espera claves como 'column', 'groups', 'grouping_mode',
                       'primary_vi_name', 'fixed_vi_name', 'fixed_descriptor_display'.
        :return: Diccionario {group_key_from_config: [normalized_data_array1, ...], ...}
                 Retorna diccionario vacío si hay errores o no hay datos.
        :raises ValueError: Si la configuración es inválida o faltan datos cruciales.
        """
        logger.info(f"Iniciando preparación de datos normalizados para estudio {study_id}, config: {config}")
        normalized_data_by_selected_group = {} # {config_group_key: [array1, array2]}

        target_column_str = config.get('column')
        if not target_column_str:
            raise ValueError("No se especificó la columna para el análisis continuo.")

        # Parsear la columna objetivo (formato Atributo/Columna/Unidad)
        # Necesitamos el nombre exacto de la columna como está en el DataFrame leído por _read_processed_file_data
        # _read_processed_file_data usa la línea 3 del archivo procesado (col_names_line.split(';'))
        # y luego sanea estos nombres.
        # La columna 'Tiempo' es añadida por file_handlers.leer_seccion y debería estar presente.
        
        # Para obtener el nombre de columna que _read_processed_file_data usará:
        # 1. Leer un archivo de muestra para obtener los nombres de columna saneados.
        # 2. Mapear el target_column_str (Attr/Col/Unit) al nombre de columna saneado.
        # Esto es un poco complejo aquí. Por ahora, asumiremos que el nombre de la columna
        # en el DataFrame es la parte 'Columna' del target_column_str.
        # Y que 'Tiempo' es el nombre de la columna de tiempo.
        try:
            parsed_target_col_parts = target_column_str.split('/', 2)
            data_col_name_in_df = parsed_target_col_parts[1] # Asumir que esta es la parte 'Columna'
            time_col_name_in_df = "Tiempo" # Nombre estándar de la columna de tiempo
        except IndexError:
            raise ValueError(f"Formato de columna inválido: '{target_column_str}'. Debe ser Atributo/Columna/Unidad.")

        logger.debug(f"Columna de datos objetivo en DataFrame: '{data_col_name_in_df}', Columna de tiempo: '{time_col_name_in_df}'")

        # Obtener todos los archivos procesados de Cinemática y su mapeo a claves de grupo completas
        # _identify_study_groups devuelve {file_base_key: full_group_key}
        all_files_to_full_group_keys_map, all_unique_full_group_keys_set = self._identify_study_groups(study_id, "Cinematica")

        # 1. Get all processed cinematic files once
        processed_files_info, _ = self.file_service.get_study_files(
            study_id=study_id, page=1, per_page=10000,
            file_type='Processed', frequency="Cinematica"
        )

        # 2. Create a map from file_base (e.g., "P01_CMJ_PRE_1") to its full Path object
        processed_path_by_file_base = {}
        for pf_info in processed_files_info:
            # Derive file_base from the processed file's path stem (e.g., "P01_CMJ_PRE_1_Cinematica" -> "P01_CMJ_PRE_1")
            file_stem = pf_info['path'].stem
            if file_stem.endswith("_Cinematica"): # Assuming frequency is appended like this
                base_name_for_map = file_stem[:-len("_Cinematica")]
                processed_path_by_file_base[base_name_for_map] = pf_info['path']
            else:
                logger.warning(f"Processed file stem '{file_stem}' does not end with '_Cinematica'. Cannot map reliably.")

        # 3. Populate files_by_full_group_key using the maps
        files_by_full_group_key = {}
        for file_base_from_study_groups, full_key in all_files_to_full_group_keys_map.items():
            # file_base_from_study_groups is the key from _identify_study_groups (e.g., "P01_CMJ_PRE_1")
            if full_key not in files_by_full_group_key:
                files_by_full_group_key[full_key] = []
            
            found_path = processed_path_by_file_base.get(file_base_from_study_groups)
            if found_path:
                files_by_full_group_key[full_key].append(found_path)
            else:
                logger.warning(f"No se encontró la ruta del archivo procesado para el archivo base '{file_base_from_study_groups}' (grupo '{full_key}') usando el mapa preconstruido.")

        selected_group_keys_from_config = config.get('groups', []) # Estas son las claves que el usuario seleccionó
        grouping_mode = config.get('grouping_mode')
        
        for selected_key_from_ui in selected_group_keys_from_config:
            # selected_key_from_ui es la clave original del grupo tal como se seleccionó en la UI.
            # En modo 1VI, esta es una clave parcial (ej: "Condicion=PRE").
            # En modo 2VIs, esta es una clave completa (ej: "Condicion=PRE;Salto=CMJ").
            # El display name usado en la UI se deriva de esta clave.
            
            normalized_data_for_this_selected_group = []
            
            # Identificar las claves de tabla completas que corresponden a este selected_key_from_ui
            # Y determinar la 'effective_group_key' que se usará como clave en normalized_data_by_selected_group
            full_keys_to_load_for_this_group_set = set()
            effective_group_key_for_dict = selected_key_from_ui

            if grouping_mode == "1VI":
                primary_vi_name = config.get('primary_vi_name')
                if not primary_vi_name:
                    logger.error("Modo 1VI seleccionado pero primary_vi_name no está en la configuración.")
                    normalized_data_by_selected_group[selected_key_from_ui] = []
                    continue
                
                # selected_key_from_ui es una CLAVE COMPLETA (e.g., "Edad=Joven;Peso=OS")
                # Necesitamos derivar la parte relevante (e.g., "Edad=Joven")
                derived_partial_key = None
                for part in selected_key_from_ui.split(';'):
                    if part.startswith(f"{primary_vi_name}="):
                        derived_partial_key = part
                        break
                
                if not derived_partial_key:
                    logger.warning(f"En modo 1VI, no se pudo derivar la clave parcial para VI '{primary_vi_name}' a partir de la clave completa seleccionada '{selected_key_from_ui}'.")
                    normalized_data_by_selected_group[selected_key_from_ui] = [] # Usar la clave completa como fallback
                    continue
                
                effective_group_key_for_dict = derived_partial_key # Usar la clave parcial como clave del diccionario de resultados
                logger.debug(f"Modo 1VI: Clave parcial derivada '{derived_partial_key}' para VI primaria '{primary_vi_name}'.")

                # Encontrar todas las claves completas en el estudio que contienen esta clave parcial derivada
                for full_key_in_study in all_unique_full_group_keys_set:
                    if derived_partial_key in full_key_in_study.split(';'):
                        full_keys_to_load_for_this_group_set.add(full_key_in_study)
            
            elif grouping_mode == "2VIs":
                # selected_key_from_ui es una clave completa y es la que se usa.
                full_keys_to_load_for_this_group_set.add(selected_key_from_ui)
                # effective_group_key_for_dict ya es selected_key_from_ui
            else: # Modo desconocido o no especificado
                logger.warning(f"Modo de agrupación desconocido o no especificado: '{grouping_mode}'. Tratando clave '{selected_key_from_ui}' como completa.")
                full_keys_to_load_for_this_group_set.add(selected_key_from_ui)
                # effective_group_key_for_dict ya es selected_key_from_ui

            if not full_keys_to_load_for_this_group_set:
                logger.warning(f"No se encontraron claves de archivo completas para el grupo UI '{selected_key_from_ui}' (clave efectiva '{effective_group_key_for_dict}') en modo '{grouping_mode}'.")
                normalized_data_by_selected_group[effective_group_key_for_dict] = []
                continue

            logger.debug(f"Para el grupo UI '{selected_key_from_ui}' (clave efectiva '{effective_group_key_for_dict}'), se cargarán datos de las siguientes claves completas: {full_keys_to_load_for_this_group_set}")

            # Si la clave efectiva no está aún en el dict de resultados, inicializarla
            if effective_group_key_for_dict not in normalized_data_by_selected_group:
                normalized_data_by_selected_group[effective_group_key_for_dict] = []

            for full_key_to_process in full_keys_to_load_for_this_group_set:
                file_paths_for_full_key = files_by_full_group_key.get(full_key_to_process, [])
                if not file_paths_for_full_key:
                    logger.warning(f"No se encontraron rutas de archivo para la clave completa '{full_key_to_process}' (derivada del grupo UI '{selected_key_from_ui}', clave efectiva '{effective_group_key_for_dict}').")
                    continue

                for file_path in file_paths_for_full_key:
                    logger.debug(f"Procesando archivo {file_path} para normalización (grupo UI '{selected_key_from_ui}', clave completa '{full_key_to_process}')")
                    df = self._read_processed_file_data(file_path)
                    if df is None or df.empty:
                        logger.warning(f"No se pudieron leer datos del archivo {file_path.name}.")
                        continue

                    if data_col_name_in_df not in df.columns:
                        logger.warning(f"Columna de datos '{data_col_name_in_df}' no encontrada en {file_path.name}. Columnas disponibles: {df.columns.tolist()}")
                        continue
                    if time_col_name_in_df not in df.columns:
                        logger.warning(f"Columna de tiempo '{time_col_name_in_df}' no encontrada en {file_path.name}. Columnas disponibles: {df.columns.tolist()}")
                        continue
                    
                    data_series = df[data_col_name_in_df]
                    time_series_str = df[time_col_name_in_df]

                    # Convertir tiempo a numérico, manejando errores
                    try:
                        time_series_numeric = pd.to_numeric(time_series_str, errors='raise')
                    except ValueError as e:
                        logger.error(f"Error convirtiendo columna de tiempo a numérico en {file_path.name}: {e}. Se omite este archivo.")
                        continue
                    
                    # Asegurar que no haya NaNs en tiempo o datos que puedan causar problemas
                    valid_indices = ~time_series_numeric.isnull() & ~data_series.isnull()
                    if not valid_indices.any():
                        logger.warning(f"No hay datos válidos (no-NaN) para tiempo/datos en {file_path.name} después de filtrar NaNs. Se omite.")
                        continue
                    
                    time_series_clean = time_series_numeric[valid_indices]
                    data_series_clean = data_series[valid_indices]

                    if len(time_series_clean) < 2:
                        logger.warning(f"No hay suficientes puntos de datos (<2) después de limpiar NaNs en {file_path.name} para interpolación. Se omite.")
                        continue

                    try:
                        normalized_array = processors.normalize_temporal_data(
                            data_series=data_series_clean,
                            time_series=time_series_clean
                            # target_points es 101 por defecto en la función
                        )
                        # Añadir a la lista de la clave efectiva (que podría ser parcial en 1VI)
                        normalized_data_by_selected_group[effective_group_key_for_dict].append(normalized_array)
                        logger.debug(f"Datos de {file_path.name} normalizados a {len(normalized_array)} puntos y añadidos al grupo efectivo '{effective_group_key_for_dict}'.")
                    except ValueError as e_norm:
                        logger.error(f"Error normalizando datos de {file_path.name} para grupo efectivo '{effective_group_key_for_dict}': {e_norm}")
                    except Exception as e_gen_norm:
                        logger.error(f"Error general durante normalización de {file_path.name} para grupo efectivo '{effective_group_key_for_dict}': {e_gen_norm}", exc_info=True)
            
            # Log después de procesar todas las claves completas para un grupo UI/efectivo
            count_for_effective_key = len(normalized_data_by_selected_group.get(effective_group_key_for_dict, []))
            logger.info(f"Para el grupo efectivo '{effective_group_key_for_dict}' (originado de UI '{selected_key_from_ui}'), se normalizaron {count_for_effective_key} series de datos.")

        return normalized_data_by_selected_group


    def perform_continuous_analysis(self, study_id: int, config: dict):
        """
        Realiza un análisis continuo (SPM) basado en la configuración proporcionada.
        Por ahora, solo prepara los datos normalizados y los loguea.

        :param study_id: ID del estudio.
        :param config: Diccionario con la configuración del análisis continuo.
        :return: (Temporal) Diccionario con los datos normalizados.
                 Futuro: Resultados del análisis SPM, rutas a gráficos, etc.
        :raises ValueError: Si la configuración es inválida o faltan datos.
        """
        analysis_name_log = config.get('analysis_name', 'N/A')
        logger.info(f"Inicio perform_continuous_analysis para estudio {study_id}: {analysis_name_log}")
        logger.debug(f"Configuración recibida para análisis continuo: {config}")

        # Initialize results_payload with default/error status
        results_payload = {
            "status": "error",
            "message": "Análisis continuo no completado.",
            "normalized_data_summary": None,
            "spm_results": None, # This will store the dict form of spm_inference
            "config_path": None,
            "spm_results_path": None,
            "continuous_plot_path": None, # Path to static PNG
            "continuous_interactive_plot_path": None, # Path to interactive HTML
            "output_dir": None
        }
        spm_inference = None # To store the spm1d inference object

        try:
            normalized_data = self._get_normalized_data_for_groups(study_id, config)

            if not normalized_data or not any(normalized_data.values()):
                logger.warning("No se generaron datos normalizados para el análisis continuo.")
                results_payload["message"] = "No se pudieron normalizar datos."
                return results_payload # Early exit

            results_payload["normalized_data_summary"] = {
                grp: {"count": len(arrays), "shape_first": arrays[0].shape if arrays else None}
                for grp, arrays in normalized_data.items()
            }

            logger.info("Datos normalizados preparados:")
            group_keys_for_spm = list(normalized_data.keys())
            # spm_test_results = None # Replaced by spm_inference
            spm_error_message = None

            if not spm1d:
                spm_error_message = "La librería spm1d no está instalada. No se puede realizar el análisis SPM."
                logger.error(spm_error_message)
            elif len(group_keys_for_spm) < 1:
                spm_error_message = "No hay grupos con datos normalizados para el análisis SPM."
                logger.error(spm_error_message)
            elif len(group_keys_for_spm) == 1:
                # Podríamos hacer un t-test one-sample si tuviera sentido (comparar contra 0)
                # Por ahora, asumimos que se necesitan al menos 2 grupos para comparar.
                spm_error_message = "Se necesita al menos dos grupos para realizar una comparación SPM."
                logger.warning(spm_error_message)
            else:
                # Preparar datos para spm1d: lista de listas/arrays de NumPy
                # La estructura de normalized_data es {group_key: [array1, array2, ...]}
                # spm1d.stats.ttest([Y0a, Y0b,...], [Y1a, Y1b,...])
                # spm1d.stats.anova1([Y0a, Y0b,...], [Y1a, Y1b,...], [Y2a, Y2b,...])
                
                data_for_spm = [normalized_data[key] for key in group_keys_for_spm if normalized_data[key]] # Asegurar que no haya listas vacías
                
                # Filtrar grupos que no tengan datos suficientes (al menos una observación)
                # Convert each group's list of 1D arrays into a 2D NumPy array for spm1d
                valid_data_for_spm = [np.array(group_data) for group_data in data_for_spm if group_data and len(group_data) > 0]
                if len(valid_data_for_spm) < 2:
                    spm_error_message = "Se necesitan al menos dos grupos con observaciones válidas para el análisis SPM."
                    logger.error(spm_error_message)
                else:
                    try:
                        if len(valid_data_for_spm) == 2:
                            logger.info(f"Realizando spm1d.stats.ttest2 para grupos: {group_keys_for_spm}")
                            # Asumir t-test independiente por ahora. Paired t-test (ttest_paired) necesitaría
                            # que los datos estén alineados por sujeto, lo cual no está garantizado por _get_normalized_data_for_groups.
                            # ttest2 is for two independent samples. It does not take equal_var.
                            spm_inference = spm1d.stats.ttest2(*valid_data_for_spm)
                            logger.info("Resultado t-test2 SPM (primeros 10 puntos del estadístico t):")
                            logger.info(spm_inference.z[:10])
                            # The spm_test_results dictionary will be built from spm_inference later, before saving.
                        
                        elif len(valid_data_for_spm) > 2:
                            logger.info(f"Realizando spm1d.stats.anova1 para {len(valid_data_for_spm)} grupos: {group_keys_for_spm}")
                            # ANOVA de un factor para muestras independientes
                            # Pass valid_data_for_spm as a list directly.
                            # Explicitly set equal_var=True to use standard ANOVA df calculation
                            # and avoid potential issues in spm1d's Welch ANOVA path (REML df estimation).
                            spm_inference = spm1d.stats.anova1(valid_data_for_spm, equal_var=True)
                            logger.info("Resultado ANOVA SPM (primeros 10 puntos del estadístico F):")
                            logger.info(spm_inference.z[:10]) # .z es la curva F
                            # The spm_test_results dictionary will be built from spm_inference later.
                        
                        # Logging about SPM completion will be handled after inference, inside the saving block.
                        # if spm_inference: # Check if spm_inference object exists and has p-value
                        #      logger.info(f"Análisis SPM completado. P-valor global: {spm_inference.p if hasattr(spm_inference, 'p') else 'N/A'}")
                        #      # Aquí se podrían inferir clusters, etc.
                        #      # spm_inference.plot() # Para visualización directa si se ejecuta interactivamente
                        #      # spm_inference.print_results()

                    except Exception as e_spm:
                        spm_error_message = f"Error durante la ejecución del análisis SPM: {e_spm}"
                        logger.error(spm_error_message, exc_info=True)
                        spm_inference = None # Ensure spm_inference is None on error

            # --- Guardar Configuración y Resultados SPM (MOVED INSIDE MAIN TRY) ---
            # Crear directorio para este análisis continuo
            analysis_output_dir_path_obj = None # Path object
            # The line `analysis_output_dir_path_obj = None # Path object` is already present from the previous patch.
            # New saving logic starts here, inside the main try block.

            try: # Nested try for saving operations
                study_path = self.file_service._get_study_path(study_id)
                if not study_path:
                    raise ValueError(f"No se pudo obtener la ruta del estudio {study_id} para guardar el análisis continuo.")

                analysis_name = config.get('analysis_name')
                if not analysis_name: 
                    raise ValueError("El nombre del análisis es requerido para guardar los resultados.")

                # --- Nueva estructura de carpetas para Análisis Continuo ---
                variable_analizada_full = config.get('column', 'VariableDesconocida')
                # Extraer "Atributo/Columna" y reemplazar '/' por un espacio o '_' para nombre de carpeta
                parts = variable_analizada_full.split('/')
                variable_folder_name_parts = parts[:2] # Tomar Atributo y Columna
                variable_folder_name = " ".join(variable_folder_name_parts).replace("/", "_") # Reemplazar / por _ si aún existe
                
                # Define the actual output directory path for this saving operation
                current_analysis_output_dir = study_path / "Analisis Continuo" / variable_folder_name / analysis_name
                current_analysis_output_dir.mkdir(parents=True, exist_ok=True)
                logger.info(f"Directorio de salida para análisis continuo '{analysis_name}' (variable '{variable_folder_name}'): {current_analysis_output_dir}")
                results_payload["output_dir"] = str(current_analysis_output_dir)

                config_file_path = current_analysis_output_dir / "config_continuous.json"
                
                # Prepare config_to_save for JSON output, potentially modifying 'groups' for 1VI mode
                config_to_save_for_json = config.copy() 
                if config_to_save_for_json.get('grouping_mode') == "1VI" and normalized_data:
                    # normalized_data keys are the effective group keys (e.g., "VI_Primaria=Descriptor")
                    effective_groups_for_json = list(normalized_data.keys())
                    config_to_save_for_json['groups'] = effective_groups_for_json
                    logger.debug(f"Ajustados 'groups' en config JSON para modo 1VI a: {effective_groups_for_json}")

                with open(config_file_path, 'w', encoding='utf-8') as f_cfg:
                    json.dump(config_to_save_for_json, f_cfg, indent=4, allow_nan=True) # Save the modified copy
                results_payload["config_path"] = str(config_file_path)
                logger.info(f"Configuración del análisis continuo guardada en: {config_file_path}")

                if spm_inference: # Check the spm1d inference object
                    alpha = 0.05 # Standard alpha level for inference
                    inference_performed_successfully = False
                    try:
                        # Determine test type to call inference appropriately
                        # This re-evaluation is to ensure scope, similar to original logic for test_type
                        data_for_spm_check = [normalized_data[key] for key in group_keys_for_spm if normalized_data[key]]
                        valid_data_for_spm_check = [gd for gd in data_for_spm_check if gd and len(gd) > 0]

                        if len(valid_data_for_spm_check) > 2: # ANOVA
                            spm_inference.inference(alpha) # Modifies in-place
                            logger.info(f"SPM .inference() llamada para ANOVA con alpha={alpha}.")
                        elif len(valid_data_for_spm_check) == 2: # t-test
                            spm_inference.inference(alpha, two_tailed=True) # Modifies in-place
                            logger.info(f"SPM .inference() llamada para t-test (two_tailed=True) con alpha={alpha}.")
                        else:
                            logger.warning("Número de grupos válidos para SPM no es 2 o más de 2, inferencia no llamada.")
                        inference_performed_successfully = True
                    except AttributeError as ae:
                        logger.warning(f"El objeto SPM ({type(spm_inference).__name__}) puede no tener el método '.inference()' o falló. Error: {ae}")
                    except Exception as e_inf:
                        logger.error(f"Error durante la ejecución de .inference() en SPM: {e_inf}", exc_info=True)

                    current_spm_results_dict = {
                        "test_type": "unknown",
                        "alpha_level": alpha,
                        "stat_curve": spm_inference.z.tolist() if hasattr(spm_inference, 'z') else None,
                        # Convert df elements to Python int
                        "df": [int(d) for d in spm_inference.df] if hasattr(spm_inference, 'df') and spm_inference.df is not None else None,
                        "critical_threshold": spm_inference.zstar if hasattr(spm_inference, 'zstar') else None,
                        "significant_clusters_found": False,
                        "clusters": []
                    }
                    
                    if hasattr(spm_inference, 'p'): # p-value from permutations, if available
                        current_spm_results_dict["global_p_value_permutations"] = spm_inference.p

                    if inference_performed_successfully and hasattr(spm_inference, 'clusters') and spm_inference.clusters:
                        current_spm_results_dict["significant_clusters_found"] = True
                        for clus in spm_inference.clusters:
                            current_spm_results_dict["clusters"].append({
                                # Convert cluster node coordinates to Python int
                                "start_node": int(clus.x0),
                                "end_node": int(clus.x1),
                                "peak_node": int(clus.peak_coord),
                                "peak_value": float(clus.peak_value), # Ensure float for consistency
                                "mass": float(clus.mass),             # Ensure float
                                "p_value": float(clus.P)              # Ensure float
                            })
                        logger.info(f"Resultados de inferencia SPM: {len(spm_inference.clusters)} clúster(es) significativos encontrados.")
                    elif inference_performed_successfully:
                        logger.info("Resultados de inferencia SPM: No se encontraron clústeres significativos.")
                    else:
                        logger.info("Inferencia SPM no realizada o falló. Resultados básicos guardados.")

                    # Determine test_type (consistent with inference call)
                    if len(valid_data_for_spm_check) == 2: current_spm_results_dict["test_type"] = "ttest2"
                    elif len(valid_data_for_spm_check) > 2: current_spm_results_dict["test_type"] = "anova1"
                    
                    results_payload["spm_results"] = current_spm_results_dict

                    spm_results_file_path = current_analysis_output_dir / "spm_results.json"
                    with open(spm_results_file_path, 'w', encoding='utf-8') as f_spm:
                        json.dump(current_spm_results_dict, f_spm, indent=4, allow_nan=True)
                    results_payload["spm_results_path"] = str(spm_results_file_path)
                    logger.info(f"Resultados SPM guardados en: {spm_results_file_path}")

                    # --- Generar Gráfico SPM ---
                    try:
                        study_aliases = self.study_service.get_study_aliases(study_id)
                        group_display_names_for_plot = []
                        # group_keys_for_spm are the keys of normalized_data
                        for group_key in group_keys_for_spm: 
                            parts = []
                            for item in group_key.split(';'):
                                vi_name, desc_value = item.split('=', 1)
                                alias = study_aliases.get(desc_value, desc_value)
                                parts.append(f"{vi_name}: {alias}")
                            group_display_names_for_plot.append(", ".join(parts))
                        
                        # Variable name for plot (use the full string as requested)
                        plot_variable_name = config.get('column', 'N/A')

                        # Prepare parameters for plotting, including display options from config
                        plot_params_for_charting = current_spm_results_dict.copy()
                        plot_params_for_charting['show_std_dev'] = config.get('show_std_dev', False)
                        plot_params_for_charting['show_conf_int'] = config.get('show_conf_int', False)
                        plot_params_for_charting['show_sem'] = config.get('show_sem', False)
                        
                        # Nuevas opciones de anotación y delimitación
                        plot_params_for_charting['annotate_spm_clusters_bottom'] = config.get('annotate_spm_clusters_bottom', True)
                        plot_params_for_charting['annotate_spm_range_top'] = config.get('annotate_spm_range_top', True)
                        plot_params_for_charting['delimit_time_range'] = config.get('delimit_time_range', False)
                        plot_params_for_charting['time_min'] = config.get('time_min', 0.0)
                        plot_params_for_charting['time_max'] = config.get('time_max', 100.0)
                        plot_params_for_charting['show_full_time_with_delimiters'] = config.get('show_full_time_with_delimiters', True)
                        plot_params_for_charting['add_time_range_label'] = config.get('add_time_range_label', False)
                        plot_params_for_charting['time_range_label_text'] = config.get('time_range_label_text', '')

                        spm_plot_path = current_analysis_output_dir / "spm_plot.png"
                        charting.create_spm_results_plot(
                            normalized_data_by_group=normalized_data, # The actual data
                            spm_results=plot_params_for_charting,     # Augmented dict with display flags
                            group_legend_names=group_display_names_for_plot,
                            variable_name=plot_variable_name, # Full column name
                            output_path=spm_plot_path
                        )
                        results_payload["continuous_plot_path"] = str(spm_plot_path)
                        logger.info(f"Gráfico de análisis continuo guardado en: {spm_plot_path}")

                        # --- Generar Gráfico SPM Interactivo (Condicional) ---
                        if config.get("generate_interactive_plot", True): # Default to True if key missing
                            spm_interactive_plot_path = current_analysis_output_dir / "spm_plot_interactive.html"
                            try:
                                charting.create_interactive_spm_results_plot(
                                    normalized_data_by_group=normalized_data,
                                    spm_results=plot_params_for_charting, # Same params as static plot
                                    group_legend_names=group_display_names_for_plot,
                                    variable_name=plot_variable_name,
                                    output_path=spm_interactive_plot_path
                                )
                                results_payload["continuous_interactive_plot_path"] = str(spm_interactive_plot_path)
                                logger.info(f"Gráfico SPM interactivo guardado en: {spm_interactive_plot_path}")
                            except Exception as e_iplot:
                                logger.error(f"Error generando gráfico SPM interactivo: {e_iplot}", exc_info=True)
                                results_payload["continuous_interactive_plot_path"] = None
                        else:
                            logger.info("Generación de gráfico SPM interactivo omitida según configuración.")
                            results_payload["continuous_interactive_plot_path"] = None
                    
                    except Exception as e_plot: # Catch errors from static plot generation or other issues
                        logger.error(f"Error generando gráfico estático de análisis continuo o procesando: {e_plot}", exc_info=True)
                        results_payload["continuous_plot_path"] = None
                        results_payload["continuous_interactive_plot_path"] = None
                
                # Update final status and message based on successful saving and analysis
                if spm_inference and not spm_error_message:
                    results_payload["status"] = "success"
                    results_payload["message"] = "Análisis SPM y guardado completados."
                elif not spm_error_message: # Normalization OK, SPM not run or failed, but saving config OK
                    results_payload["status"] = "partial_success"
                    results_payload["message"] = "Datos normalizados generados y configuración guardada. Análisis SPM no ejecutado o falló."
                else: # spm_error_message exists
                    results_payload["status"] = "error_in_spm_analysis" 
                    results_payload["message"] = f"Configuración guardada. Error en análisis SPM: {spm_error_message}"

            except Exception as e_save: # Catch errors during saving
                logger.error(f"Error guardando resultados del análisis continuo '{analysis_name_log}': {e_save}", exc_info=True)
                results_payload["status"] = "error_saving"
                # Append to existing message if analysis had a message, otherwise set new one
                base_message = spm_error_message if spm_error_message else "Error en la fase de análisis."
                # Use the message from the results_payload if it's more specific than the default
                if results_payload.get("message") and results_payload["message"] != "Análisis continuo no completado.":
                     base_message = results_payload["message"]
                results_payload["message"] = f"{base_message} Adicionalmente, error al guardar resultados: {e_save}."
            
            # This return is now the main successful return path from the primary try block
            return results_payload

        except ValueError as ve: # Outer handler for normalization/validation errors
            logger.error(f"Error de validación o datos al preparar análisis continuo '{analysis_name_log}': {ve}", exc_info=True)
            results_payload["message"] = str(ve) 
            # results_payload["status"] is already "error" by default
            return results_payload 
        except Exception as e: # Outer catch-all for unexpected errors
            logger.critical(f"Error inesperado durante análisis continuo '{analysis_name_log}': {e}", exc_info=True)
            results_payload["message"] = f"Error inesperado: {e}" 
            # results_payload["status"] is already "error" by default
            return results_payload

    # --- Métodos para Gestión de Análisis Continuo ---

    def _get_continuous_analysis_base_dir(self, study_id: int) -> Path | None:
        """Obtiene el directorio base para los análisis continuos."""
        study_path = self.file_service._get_study_path(study_id)
        if not study_path:
            logger.error(f"No se pudo encontrar ruta estudio {study_id} para análisis continuo.")
            return None
        return study_path / "Analisis Continuo"

    def does_continuous_analysis_exist(self, study_id: int, variable_folder_name: str, analysis_name: str) -> bool:
        """
        Verifica si ya existe un análisis continuo con el nombre dado para una variable específica.

        :param study_id: ID del estudio.
        :param variable_folder_name: Nombre de la carpeta de la variable (ej: "LAnkleAngles X").
        :param analysis_name: Nombre del análisis a verificar.
        :return: True si existe, False en caso contrario.
        """
        base_dir = self._get_continuous_analysis_base_dir(study_id)
        if not base_dir or not base_dir.exists():
            return False
        
        # Sanitize analysis_name just in case, though it should be clean from dialog
        clean_analysis_name = analysis_name.strip()
        invalid_chars = r'<>:"/\|?*'
        if any(char in clean_analysis_name for char in invalid_chars):
            logger.warning(f"Intento de verificar existencia de análisis con nombre inválido: '{clean_analysis_name}'")
            return False # Or raise error, but for existence check, False is safer.

        analysis_path = base_dir / variable_folder_name / clean_analysis_name
        return analysis_path.exists() and analysis_path.is_dir()

    def list_continuous_analyses(self, study_id: int) -> list[dict]:
        """
        Lista los análisis continuos guardados para un estudio.

        :param study_id: ID del estudio.
        :return: Lista de diccionarios:
                 {'name': str, 'path': Path, 'config': dict, 'mtime': float,
                  'plot_path': Path | None, 'interactive_plot_path': Path | None,
                  'spm_results_path': Path | None}
        """
        analyses = []
        base_dir = self._get_continuous_analysis_base_dir(study_id)
        if not base_dir or not base_dir.exists():
            return []

        # Iterate through variable-specific subfolders (e.g., "LAnkleMoment X")
        for variable_folder_path in base_dir.iterdir():
            if not variable_folder_path.is_dir():
                continue
            
            # Now iterate through analysis name folders within each variable folder
            for analysis_specific_folder_path in variable_folder_path.iterdir():
                if analysis_specific_folder_path.is_dir():
                    analysis_name = analysis_specific_folder_path.name
                    config_path = analysis_specific_folder_path / "config_continuous.json"
                    plot_path = analysis_specific_folder_path / "spm_plot.png"
                    interactive_plot_path = analysis_specific_folder_path / "spm_plot_interactive.html"
                    spm_results_path = analysis_specific_folder_path / "spm_results.json"

                    if config_path.exists() and config_path.is_file():
                        try:
                            with open(config_path, 'r', encoding='utf-8') as f:
                                config_data = json.load(f)
                            mtime = config_path.stat().st_mtime
                            analyses.append({
                                'name': analysis_name,
                                'path': analysis_specific_folder_path, # Path to the specific analysis folder
                                'config': config_data,
                                'mtime': mtime,
                                'plot_path': plot_path if plot_path.exists() else None,
                                'interactive_plot_path': interactive_plot_path if interactive_plot_path.exists() else None,
                                'spm_results_path': spm_results_path if spm_results_path.exists() else None,
                                'config_path': config_path
                            })
                        except json.JSONDecodeError:
                            logger.error(f"Error leyendo config_continuous.json para análisis '{analysis_name}' en {analysis_specific_folder_path}.")
                        except Exception as e:
                            logger.error(f"Error procesando análisis continuo '{analysis_name}' en {analysis_specific_folder_path}: {e}", exc_info=True)
                    elif not config_path.exists() or not config_path.is_file():
                        logger.warning(f"Directorio análisis continuo '{analysis_name}' en {analysis_specific_folder_path} encontrado sin config_continuous.json válido.")
        
        analyses.sort(key=lambda x: x['mtime'], reverse=True)
        return analyses

    def _get_contributing_full_keys(self, study_id: int, frequency: str, required_parts: list[str]) -> list[str]:
        """
        Finds all unique full group keys in the study that contain ALL specified required_parts.
        A full group key is like "VI1=DescA;VI2=DescB;VI3=DescC".
        required_parts is a list like ["VI1=DescA", "VI3=DescC"].
        """
        if not required_parts:
            return []
        
        try:
            # _, all_unique_full_keys_set = self._identify_study_groups(study_id, frequency)
            # _identify_study_groups returns map and set. We need the set of full keys.
            files_to_groups_map, unique_full_keys_set = self._identify_study_groups(study_id, frequency)
            if not unique_full_keys_set:
                return []

            matching_full_keys = []
            for full_key_in_study in unique_full_keys_set:
                parts_in_study_key = set(full_key_in_study.split(';'))
                if all(req_part in parts_in_study_key for req_part in required_parts):
                    matching_full_keys.append(full_key_in_study)
            
            return sorted(list(set(matching_full_keys))) # Return sorted unique list
        except Exception as e:
            logger.error(f"Error en _get_contributing_full_keys para estudio {study_id}, partes {required_parts}: {e}", exc_info=True)
            return []

    def _delete_continuous_analysis_no_backup(self, analysis_folder_to_delete: Path):
        """
        Core logic to delete a continuous analysis folder. No backup trigger here.
        """
        # Backup is now handled by the public method or batch operation
        logger.info(f"Core logic: eliminar análisis continuo en: {analysis_folder_to_delete}")

        if not analysis_folder_to_delete.exists():
            raise FileNotFoundError(f"El directorio del análisis continuo no existe: {analysis_folder_to_delete}")
        if not analysis_folder_to_delete.is_dir():
            raise ValueError(f"La ruta del análisis continuo no es un directorio: {analysis_folder_to_delete}")

        try:
            shutil.rmtree(analysis_folder_to_delete)
            logger.info(f"Análisis continuo eliminado: {analysis_folder_to_delete}")

            # Opcional: Limpiar directorio padre (variable_folder_name) si queda vacío.
            parent_dir = analysis_folder_to_delete.parent
            if parent_dir.exists() and parent_dir.is_dir() and not any(parent_dir.iterdir()):
                # Verificar que el nombre del directorio padre no sea "Analisis Continuo" directamente
                if parent_dir.name != "Analisis Continuo":
                    logger.info(f"Directorio de variable '{parent_dir.name}' está vacío tras eliminar '{analysis_folder_to_delete.name}', eliminándolo.")
                    try:
                        parent_dir.rmdir()
                        logger.info(f"Directorio de variable '{parent_dir.name}' eliminado.")
                    except OSError as e_rmparent:
                        logger.error(f"Error eliminando directorio padre vacío '{parent_dir.name}': {e_rmparent}", exc_info=True)
                else:
                    logger.debug(f"Directorio padre es '{parent_dir.name}', no se eliminará aunque esté vacío.")
            elif parent_dir.exists() and parent_dir.is_dir():
                logger.debug(f"Directorio de variable '{parent_dir.name}' no está vacío, no se eliminará.")

        except OSError as e:
            logger.error(f"Error eliminando directorio análisis continuo {analysis_folder_to_delete}: {e}", exc_info=True)
            raise

    def delete_continuous_analysis(self, analysis_folder_to_delete: Path):
        """
        Elimina la carpeta y contenido de un análisis continuo específico, triggering backup.
        """
        try:
            create_backup(backup_type='automatic')
        except Exception as e_backup:
            logger.error(f"Error creating automatic backup before deleting continuous analysis {analysis_folder_to_delete.name}: {e_backup}", exc_info=True)
            # Log and continue for now
        
        if self.undo_manager.is_undo_enabled():
            if not self.undo_manager.prepare_undo_cache():
                logger.error(f"Failed to prepare undo cache for deleting continuous analysis {analysis_folder_to_delete}. Aborting delete operation.")
                raise Exception(f"Failed to prepare undo cache for deleting continuous analysis {analysis_folder_to_delete}. Deletion aborted.")

            if not self.undo_manager.cache_item_for_undo(str(analysis_folder_to_delete), "continuous_analysis_folder"):
                logger.warning(f"Failed to cache continuous analysis folder {analysis_folder_to_delete} for undo. Deletion will proceed but undo might be partial.")
        
        try:
            self._delete_continuous_analysis_no_backup(analysis_folder_to_delete)
        except Exception as e:
            logger.error(f"Error during deletion of continuous analysis {analysis_folder_to_delete} after undo preparation: {e}", exc_info=True)
            raise

    # --- Métodos para Análisis Discreto (Fase 6) ---

    def get_discrete_analysis_tables_path(self, study_id: int) -> Path | None:
        """
        Obtiene la ruta base donde se guardan las tablas de resumen discreto.

        :param study_id: ID del estudio.
        :return: Path al directorio de tablas o None si no se puede determinar.
        """
        try:
            study_path = self.file_service._get_study_path(study_id)
            if not study_path:
                logger.warning(f"No se pudo obtener ruta estudio {study_id} "
                               f"para buscar tablas discretas.")
                return None
            # Asume estructura definida en generate_discrete_summary_tables
            # Nota: Podría ser más robusto guardando ruta relativa en config.
            tables_path = study_path / "Analisis Discreto" / "Tablas"
            return tables_path
        except Exception as e:
            logger.error(f"Error obteniendo ruta tablas discretas para "
                         f"estudio {study_id}: {e}", exc_info=True)
            return None

    def _delete_discrete_summary_table_no_backup(self, table_path_str: str):
        """
        Core logic to delete a discrete summary table (.xlsx and .csv). No backup trigger.
        """
        # Backup is handled by public method or batch operation
        xlsx_path = Path(table_path_str)
        if not xlsx_path.exists():
            raise FileNotFoundError(f"El archivo de tabla .xlsx no existe: {xlsx_path}")
        if not xlsx_path.is_file():
            raise ValueError(f"La ruta no es un archivo: {xlsx_path}")
        if xlsx_path.suffix.lower() != '.xlsx':
            raise ValueError(f"El archivo no parece ser una tabla .xlsx: {xlsx_path}")

        try:
            # Eliminar el archivo .xlsx
            xlsx_path.unlink()
            logger.info(f"Tabla de resumen .xlsx eliminada: {xlsx_path}")

            # Intentar eliminar el archivo .csv correspondiente
            csv_path = xlsx_path.with_suffix('.csv')
            if csv_path.exists() and csv_path.is_file():
                try:
                    csv_path.unlink()
                    logger.info(f"Tabla interna .csv correspondiente eliminada: {csv_path}")
                except OSError as e_csv:
                    logger.warning(f"No se pudo eliminar el archivo .csv interno {csv_path}: {e_csv}", exc_info=True)
            else:
                logger.debug(f"No se encontró archivo .csv interno correspondiente para {xlsx_path.name} en {csv_path.parent}, o no es un archivo.")

            # Opcional: Limpiar directorios vacíos (Tipo de Dato, etc.)
            # Esta lógica podría ser más compleja si se quiere asegurar que solo se borren
            # carpetas relacionadas con "Analisis Discreto/Tablas".
            # Por ahora, se omite para mantener el foco en la eliminación de archivos.

        except OSError as e:
            logger.error(f"Error al eliminar la tabla {xlsx_path}: {e}", exc_info=True)
            raise

    def delete_discrete_summary_table(self, table_path_str: str):
        """
        Elimina un archivo de tabla de resumen discreto específico (.xlsx) y su
        correspondiente archivo .csv interno si existe, triggering backup.
        """
        try:
            create_backup(backup_type='automatic')
        except Exception as e_backup:
            logger.error(f"Error creating automatic backup before deleting discrete summary table {table_path_str}: {e_backup}", exc_info=True)
            # Log and continue for now
        
        if self.undo_manager.is_undo_enabled():
            if not self.undo_manager.prepare_undo_cache():
                logger.error(f"Failed to prepare undo cache for deleting discrete table {table_path_str}. Aborting delete operation.")
                raise Exception(f"Failed to prepare undo cache for deleting discrete table {table_path_str}. Deletion aborted.")

            # Cache the .xlsx file. The .csv is implicitly handled by DB restore if needed, or not cached if purely derived.
            # For file system undo, caching the primary user-facing file (.xlsx) is key.
            xlsx_path = Path(table_path_str)
            if not self.undo_manager.cache_item_for_undo(str(xlsx_path), "discrete_summary_table_xlsx"):
                logger.warning(f"Failed to cache discrete table {xlsx_path} for undo. Deletion will proceed but undo might be partial.")
            
            # Also cache the corresponding .csv file if it exists
            csv_path = xlsx_path.with_suffix('.csv')
            if csv_path.exists() and csv_path.is_file():
                if not self.undo_manager.cache_item_for_undo(str(csv_path), "discrete_summary_table_csv"):
                    logger.warning(f"Failed to cache corresponding CSV table {csv_path} for undo. Deletion will proceed but undo might be partial.")
        
        try:
            self._delete_discrete_summary_table_no_backup(table_path_str)
        except Exception as e:
            logger.error(f"Error during deletion of discrete table {table_path_str} after undo preparation: {e}", exc_info=True)
            raise

    def _extract_stats_from_processed_file(self, file_path: Path, calculation: str) -> list | None:
        """Lee las últimas líneas de un archivo procesado y extrae la fila de datos para el cálculo especificado."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            if len(lines) < 3:  # Necesita al menos las 3 líneas de stats
                logger.warning(f"Archivo {file_path.name} no tiene suficientes "
                               f"líneas para extraer estadísticas.")
                return None

            # Buscar la línea del cálculo en las últimas 3 líneas
            calc_line_prefix = f";;{calculation.upper()};"
            for line in reversed(lines[-3:]):
                if line.startswith(calc_line_prefix):
                    # Quitar prefijo y dividir por ';'
                    # Devolver valores como lista de strings (incluye vacíos)
                    return line.strip()[len(calc_line_prefix):].split(';')
            logger.warning(f"No se encontró línea de cálculo '{calculation}' "
                           f"en {file_path.name}")
            return None
        except Exception as e:
            logger.error(f"Error extrayendo stats de {file_path.name} para "
                         f"cálculo {calculation}: {e}", exc_info=True)
            return None

    def _parse_processed_file_headers(self, file_path: Path) -> tuple[list, list, list] | None:
        """Lee las líneas 1, 2 y 3 (atributos, columnas, unidades) de un archivo procesado."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            if len(lines) < 4:  # Necesita num_frames, attr, col, unit
                logger.warning(f"Archivo {file_path.name} no tiene suficientes "
                               f"líneas de cabecera.")
                return None
            # Líneas 1, 2, 3 (índices 1, 2, 3)
            # Omitir Frame, SubFrame, Tiempo y limpiar prefijo PteXX:
            atributos_raw = lines[1].strip().split(';')[3:]
            columnas_raw = lines[2].strip().split(';')[3:]
            unidades_raw = lines[3].strip().split(';')[3:]

            # Función auxiliar para limpiar prefijo "PteXX:"
            def clean_prefix(name: str) -> str:
                name = name.strip()
                if ':' in name:
                    # Intentar dividir y tomar la segunda parte
                    parts = name.split(':', 1)
                    if len(parts) > 1:
                        return parts[1].strip()
                # Si no hay ':' o la división falla, devolver el nombre original (limpio)
                return name

            # Limpiar prefijos de atributos y columnas
            atributos = [clean_prefix(a) for a in atributos_raw]
            columnas = [clean_prefix(c) for c in columnas_raw]
            # Limpiar espacios en unidades (no deberían tener prefijo)
            unidades = [u.strip() for u in unidades_raw]

            # Asegurar que todas las listas tengan la misma longitud
            max_len = max(len(atributos), len(columnas), len(unidades))
            atributos.extend([''] * (max_len - len(atributos)))
            columnas.extend([''] * (max_len - len(columnas)))
            unidades.extend([''] * (max_len - len(unidades)))

            return atributos, columnas, unidades
        except Exception as e:
            logger.error(f"Error parseando cabeceras de {file_path.name}: {e}",
                         exc_info=True)
            return None

    def generate_discrete_summary_tables(self, study_id: int):
        """
        Genera tablas resumen CSV para cálculos discretos (Max, Min, Rango)
        agrupados por frecuencia y combinación de sub-valores.
        Enfocado inicialmente en 'Cinematica'.

        :param study_id: ID del estudio.
        :return: Diccionario con rutas de los archivos generados o errores.
                 {'success': [path_str], 'errors': [error_msg]}
        """
        logger.info(f"Iniciando generación de tablas resumen discreto para "
                    f"estudio {study_id}")
        results = {'success': [], 'errors': []}
        target_frequency = "Cinematica"  # Enfocarse en Cinemática por ahora
        calculations = ["Maximo", "Minimo", "Rango"]

        try:
            study_path = self.file_service._get_study_path(study_id)
            if not study_path:
                results['errors'].append(f"No se pudo encontrar la ruta del estudio {study_id}.")
                return results

            study_details = self.study_service.get_study_details(study_id)
            defined_descriptors = [d.strip() for d in
                                   (study_details.get('sub-valores', '') or '')
                                   .split(',') if d.strip()]

            # 1. Encontrar y agrupar archivos procesados de Cinemática
            # { 'Desc1_Desc2': [path1, path2,...], ... }
            files_by_descriptor_combo = {}
            processed_files_info, _ = self.file_service.get_study_files( # Corregir nombre variable
                study_id=study_id,
                page=1, # Eliminar el 'page=1' duplicado
                per_page=10000, file_type='Processed', frequency=target_frequency
            )
            path_map = {f_info['path'].stem.split(f'_{target_frequency}')[0]: f_info['path']
                        for f_info in processed_files_info}

            # --- Identificar grupos y mapear archivos ---
            # Usar _identify_study_groups para obtener el mapeo archivo_base -> group_key
            try:
                files_to_groups, _ = self._identify_study_groups(study_id, target_frequency)
            except ValueError as e_group:
                results['errors'].append(f"Error identificando grupos para {target_frequency}: {e_group}")
                return results

            files_by_group_key = {} # Inicializar aquí
            for file_base_key, group_key in files_to_groups.items():
                if group_key not in files_by_group_key:
                    files_by_group_key[group_key] = []
                # Buscar la ruta completa usando el file_base_key
                full_path = path_map.get(file_base_key)
                if full_path:
                    files_by_group_key[group_key].append(full_path)
                else:
                    logger.warning(f"No se encontró la ruta completa para el archivo base '{file_base_key}' del grupo '{group_key}'.")


            if not files_by_group_key:
                results['errors'].append(f"No se encontraron archivos válidos agrupables por VIs para '{target_frequency}'.")
                return results

            # 2. Preparar directorio de salida
            output_base_dir = (study_path / "Analisis Discreto" / "Tablas" /
                               target_frequency)
            # Asegurar que el directorio exista, pero no limpiar su contenido.
            output_base_dir.mkdir(parents=True, exist_ok=True)

            # 3. Generar tabla para cada grupo y cálculo
            for group_key, file_paths in files_by_group_key.items():
                if not file_paths:
                    continue

                # Crear nombre de archivo seguro reemplazando caracteres inválidos
                safe_group_key_part = group_key.replace('=', '_').replace(';', '__')

                # Leer cabeceras desde el primer archivo del grupo
                headers = self._parse_processed_file_headers(file_paths[0])
                if not headers:
                    results['errors'].append(f"No se pudieron leer las cabeceras para el grupo '{group_key}'.")
                    continue
                atributos, columnas, unidades = headers
                # Número de columnas de datos (sin Frame, Sub, Tiempo)
                num_value_cols = len(columnas)

                # Crear MultiIndex para las columnas
                multi_index_tuples = []
                last_attr = ""
                for i in range(num_value_cols):
                    # Propagar atributo si está vacío
                    attr = atributos[i] if atributos[i] else last_attr
                    # Añadir unidad
                    multi_index_tuples.append((attr, columnas[i], unidades[i]))
                    last_attr = attr
                # Crear MultiIndex con tres niveles
                column_multi_index = pd.MultiIndex.from_tuples(
                    multi_index_tuples,
                    names=["Atributo", "Columna", "Unidad"]
                )


                for calc in calculations:
                    table_data = []
                    # Almacenar nombres base completos para el índice
                    index_names = []

                    for file_path in file_paths:
                        stats_row = self._extract_stats_from_processed_file(
                            file_path, calc
                        )
                        # Obtener nombre base completo para índice
                        base_name = file_path.stem.split(f'_{target_frequency}')[0]

                        if stats_row and len(stats_row) == num_value_cols:
                            table_data.append(stats_row)
                            index_names.append(base_name) # Guardar nombre base completo
                        else:
                            logger.warning(f"Datos de '{calc}' inconsistentes o "
                                           f"faltantes en {file_path.name}. Se "
                                           f"omitirá de las tablas resumen.")
                            # No añadir a index_names si se omite la fila

                    if table_data:
                        try:
                            # --- Crear y Guardar CSV Interno (para lógica posterior) ---
                            df_csv_internal = pd.DataFrame(table_data, columns=column_multi_index, index=index_names)
                            df_csv_internal.index.name = "ARCHIVO"

                            # Convertir CSV interno a numérico (punto decimal)
                            for col in df_csv_internal.columns:
                                if df_csv_internal[col].dtype == 'object':
                                    df_csv_internal[col] = df_csv_internal[col].str.replace(',', '.', regex=False)
                                df_csv_internal[col] = pd.to_numeric(df_csv_internal[col], errors='coerce')

                            # Usar safe_group_key_part para el nombre de archivo
                            output_csv_internal_path = output_base_dir / f"{calc}_{target_frequency}_{safe_group_key_part}.csv"
                            df_csv_internal.to_csv(output_csv_internal_path, sep=',', decimal='.',
                                                   encoding='utf-8', header=True, index=True)
                            # results['success'].append(str(output_csv_internal_path)) # CSV is internal, not counted for user
                            logger.info(f"Tabla CSV interna generada: {output_csv_internal_path}")

                            # --- Preparar DataFrame para formatos de exportación (XLSX) ---
                            # Usar los mismos datos numéricos pero con índice como columna 'ARCHIVO'
                            df_data_export = df_csv_internal.reset_index() # Mover índice 'ARCHIVO' a columna

                            # --- Construir Cabeceras Manualmente (para TSV, XLSX, SCSV) ---
                            # 4 filas de cabecera según el ejemplo TSV
                            header_rows_list = []
                            num_data_cols = len(column_multi_index)
                            total_cols = 1 + num_data_cols # ARCHIVO + datos

                            # Fila 0: Vacía (solo separadores)
                            header_rows_list.append([''] * total_cols)

                            # Fila 1: Caracteristica (Atributo) - Empieza vacía
                            row1 = [''] # Celda vacía para columna ARCHIVO
                            last_attr = None
                            for attr, _, _ in column_multi_index:
                                if attr == last_attr and last_attr is not None:
                                    row1.append('') # Simular celda combinada
                                else:
                                    row1.append(attr if attr else '')
                                    last_attr = attr
                            header_rows_list.append(row1)

                            # Fila 2: Espacio (Columna) - Empieza con 'ARCHIVO'
                            row2 = ['ARCHIVO']
                            for _, col, _ in column_multi_index:
                                row2.append(col if col else '')
                            header_rows_list.append(row2)

                            # Fila 3: Unidad - Empieza vacía
                            row3 = [''] # Celda vacía para columna ARCHIVO
                            for _, _, unit in column_multi_index:
                                row3.append(unit if unit else '')
                            header_rows_list.append(row3)

                            # --- Guardar Formatos de Exportación (usando safe_group_key_part) ---
                            # Solo se genera XLSX para el usuario. CSV es interno.

                            # Formato XLSX (Excel) - openpyxl requerido
                            if OPENPYXL_AVAILABLE:
                                output_xlsx_path = output_base_dir / f"{calc}_{target_frequency}_{safe_group_key_part}.xlsx"
                                try:
                                    # Usar ExcelWriter para acceder al objeto worksheet
                                    with pd.ExcelWriter(output_xlsx_path, engine='openpyxl') as writer:
                                        # Asegurar que la hoja 'Data' exista
                                        if not writer.book.sheetnames:
                                            ws = writer.book.create_sheet(title='Data')
                                        else:
                                            ws = writer.book.get_sheet_by_name('Data') \
                                                 if 'Data' in writer.book.sheetnames \
                                                 else writer.book.active
                                            ws.title = 'Data'

                                        # Escribir las 4 filas de cabecera manualmente
                                        for header_row_list in header_rows_list:
                                            ws.append(header_row_list)

                                        # Aplanar columnas para evitar error de pandas con MultiIndex en to_excel
                                        df_data_export_flat = df_data_export.copy()
                                        flat_column_names = [df_data_export_flat.columns[0]] # 'ARCHIVO'
                                        flat_column_names.extend([
                                            '/'.join(map(str, col_tuple))
                                            for col_tuple in df_data_export_flat.columns[1:]
                                        ])
                                        df_data_export_flat.columns = flat_column_names

                                        # Escribir datos con columnas aplanadas
                                        df_data_export_flat.to_excel(
                                            writer,
                                            sheet_name='Data',
                                            startrow=len(header_rows_list),
                                            header=False,
                                            index=False,
                                            float_format='%.4f' # Asegurar formato de float
                                        )
                                    results['success'].append(str(output_xlsx_path)) # Contar solo XLSX para el usuario
                                    logger.info(f"Tabla XLSX generada: {output_xlsx_path}")
                                except Exception as e_xlsx:
                                    error_msg = f"Error guardando XLSX {output_xlsx_path.name}: {e_xlsx}"
                                    logger.error(error_msg, exc_info=True)
                                    results['errors'].append(error_msg)
                            else:
                                logger.warning(f"Omitiendo generación XLSX para {group_key}/{calc} (openpyxl no disponible).")
                                results['errors'].append(f"openpyxl no disponible para generar {calc}_{target_frequency}_{safe_group_key_part}.xlsx")

                        except Exception as e_df:
                            # Error general al procesar este cálculo/grupo
                            error_msg = (f"Error procesando datos para "
                                         f"{calc}_{target_frequency}_{group_key}: {e_df}")
                            logger.error(error_msg, exc_info=True)
                            results['errors'].append(error_msg)
                    else:
                        # Mensaje genérico ya que se generan múltiples formatos
                        logger.warning(f"No se encontraron datos válidos para "
                                       f"generar tablas resumen para "
                                       f"{calc}_{target_frequency}_{group_key}")

        except Exception as e:
            error_msg = (f"Error inesperado durante generación tablas discretas "
                         f"para estudio {study_id}: {e}")
            logger.critical(error_msg, exc_info=True)
            results['errors'].append(error_msg)

        logger.info(f"Generación tablas discretas finalizada para estudio "
                    f"{study_id}. Éxitos: {len(results['success'])}, "
                    f"Errores: {len(results['errors'])}")
        return results

    # --- Métodos para Análisis Individual (Fase 6) ---

    def _identify_study_groups(self, study_id: int, frequency: str = "Cinematica") -> tuple[dict[str, str], set[str]]:
        """
        Identifica los grupos únicos basados en sub-valores de archivos procesados.

        :param study_id: ID del estudio.
        :param frequency: Tipo de Dato a considerar (por defecto 'Cinematica').
        :return: Tupla:
                 - Dict mapeando nombre base archivo a clave grupo (formato VI=Desc).
                 - Set de claves de grupo únicas encontradas.
        :raises ValueError: Si no se pueden obtener detalles o archivos.
        """
        logger.debug(f"Identificando grupos para estudio {study_id}, frecuencia {frequency}")
        groups_by_file_base = {}
        unique_group_keys = set()

        try:
            study_details = self.study_service.get_study_details(study_id)
            if not study_details:
                raise ValueError(f"No se pudieron obtener detalles del estudio {study_id}")
            # Obtener estructura VI
            independent_variables = study_details.get('independent_variables', [])

            processed_files, _ = self.file_service.get_study_files(
                study_id=study_id,
                page=1,
                per_page=10000,  # Obtener todos
                file_type='Processed',
                frequency=frequency
            )

            if not processed_files:
                logger.warning(f"No se encontraron archivos procesados de "
                               f"'{frequency}' para identificar grupos en "
                               f"estudio {study_id}.")
                return {}, set()

            for file_info in processed_files:
                file_path = file_info['path']
                filename = file_path.name

                # Validar nombre usando VIs y desempaquetar los 4 valores
                is_valid_name, _, extracted_descriptors, _ = validate_filename_for_study_criteria(
                    filename, independent_variables
                )
                # Solo necesitamos is_valid_name y extracted_descriptors aquí
                if not is_valid_name:
                    continue

                # Crear clave de grupo combinada (VI=Desc;...)
                try:
                    group_parts = []
                    for i, desc in enumerate(extracted_descriptors):
                        vi_name = independent_variables[i].get('name', f'VI{i+1}')
                        value = desc if desc is not None else "Nulo"
                        group_parts.append(f"{vi_name}={value}")
                    group_key = ";".join(group_parts) if group_parts else "SinGrupo"

                    # Usar nombre base sin frecuencia ni extensión como clave del dict
                    file_base_key = file_path.stem.split(f'_{frequency}')[0]

                    groups_by_file_base[file_base_key] = group_key
                    unique_group_keys.add(group_key)
                except IndexError: # Si extracted_descriptors no coincide con VIs
                    logger.warning(f"Discrepancia VIs/Sub-valores al extraer grupo de: {filename}")
                    continue

            logger.debug(f"Grupos identificados ({len(unique_group_keys)}): "
                         f"{unique_group_keys}")
            return groups_by_file_base, unique_group_keys

        except Exception as e:
            logger.error(f"Error identificando grupos para estudio {study_id}: {e}", exc_info=True)
            raise ValueError(f"Error identificando grupos: {e}")


    def get_discrete_analysis_groups(self, study_id: int, frequency: str = "Cinematica") -> list[str]:
        """
        Obtiene la lista de claves de grupos únicos para análisis discreto.

        :param study_id: ID del estudio.
        :param frequency: Tipo de Dato a considerar.
        :return: Lista ordenada de tuplas (display_name, original_key).
        """
        try:
            _, unique_group_keys = self._identify_study_groups(study_id, frequency)
            study_aliases = self.study_service.get_study_aliases(study_id)

            # Crear tuplas ("Grupo X - Display Name", original_key)
            groups_with_display_names = []
            # Ordenar claves originales para numeración consistente
            sorted_group_keys = sorted(list(unique_group_keys))
            for i, group_key in enumerate(sorted_group_keys):
                display_parts = []
                if group_key != "SinGrupo":
                    for part in group_key.split(';'):
                        vi_name, desc_value = part.split('=', 1)
                        alias = study_aliases.get(desc_value, desc_value) # Aplicar alias
                        display_parts.append(f"{vi_name}: {alias}") # Formato "VI: Alias"
                base_display_name = ", ".join(display_parts) if display_parts else "Grupo General"
                # Añadir prefijo "Grupo X - "
                full_display_name = f"Grupo {i+1} - {base_display_name}"
                groups_with_display_names.append((full_display_name, group_key)) # Guardar clave original

            # Ordenar por el nombre completo visible (ya incluye número)
            groups_with_display_names.sort()
            return groups_with_display_names

        except ValueError as e:
            logger.warning(f"No se pudieron obtener grupos para estudio {study_id}: {e}")
            return [] # Devolver vacío si hay error


    def get_filtered_discrete_analysis_groups(self, study_id: int, frequency: str, mode: str,
                                              primary_vi_name: Optional[str] = None,
                                              fixed_vi_name: Optional[str] = None,
                                              fixed_descriptor_value: Optional[str] = None) -> dict[str, str]:
        """
        Obtiene los grupos de análisis discreto FILTRADOS según el modo y selecciones.

        :param study_id: ID del estudio.
        :param frequency: Tipo de Dato seleccionada.
        :param mode: '1VI' o '2VIs'.
        :param primary_vi_name: Nombre de la VI seleccionada si mode='1VI'.
        :param fixed_vi_name: Nombre de la VI a fijar si mode='2VIs'.
        :param fixed_descriptor_value: Valor del sub-valor a fijar si mode='2VIs' (valor original, sin alias).
        :return: Diccionario {original_key: display_name} de los grupos filtrados.
                 El display_name se ajusta según el modo.
        """
        logger.info(f"Filtrando grupos: mode={mode}, freq={frequency}, primary={primary_vi_name}, fixed_vi={fixed_vi_name}, fixed_desc={fixed_descriptor_value}")
        # Obtener todos los grupos como lista de tuplas (display, key)
        all_groups_tuples = self.get_discrete_analysis_groups(study_id, frequency)
        # Convertir a dict {key: display} para facilitar búsqueda
        all_groups = {key: display for display, key in all_groups_tuples}
        aliases = self.study_service.get_study_aliases(study_id)
        filtered_groups = {}
        # temp_groups = {} # This variable was removed, causing the mismatch.

        if mode == '1VI' and primary_vi_name:
            logger.debug(f"Filtrando modo 1VI por '{primary_vi_name}'")
            # Collect unique partial keys (e.g., "Edad=Joven") and their display names
            unique_partial_keys_for_primary_vi = {} # {partial_key: display_name}
            for original_full_key, _ in all_groups.items(): # Iterar sobre todas las claves originales
                parts = original_full_key.split(';')
                for part in parts:
                    if part.startswith(f"{primary_vi_name}="):
                        try:
                            # part is like "VI_Primaria=DescriptorValor"
                            _, descriptor = part.split('=', 1)
                            alias = aliases.get(descriptor, descriptor)
                            display_name = f"{primary_vi_name}: {alias}"
                            # Store the partial key itself, ensuring uniqueness by display_name
                            # Check if display_name is already a value to avoid overwriting with a different partial_key
                            # that maps to the same display_name (unlikely with current alias logic but safe).
                            # The goal is to have a unique set of display_name -> partial_key for the UI.
                            # So, if a display_name is already mapped, we prefer the first partial_key found.
                            if display_name not in unique_partial_keys_for_primary_vi.values():
                                unique_partial_keys_for_primary_vi[part] = display_name
                            break # Found the primary VI part for this original_full_key
                        except ValueError:
                            logger.warning(f"Error parseando parte '{part}' de la clave '{original_full_key}'")
            filtered_groups = unique_partial_keys_for_primary_vi # This is now {partial_key: display_name}
            # The dialog will then use {display_name: partial_key}

        elif mode == '2VIs' and fixed_vi_name and fixed_descriptor_value:
            logger.debug(f"Filtrando modo 2VIs fijando '{fixed_vi_name}={fixed_descriptor_value}'")
            fixed_pair_str = f"{fixed_vi_name}={fixed_descriptor_value}"
            for key, _ in all_groups.items():
                parts = key.split(';')
                # Verificar si la parte fija está presente
                if fixed_pair_str in parts:
                    # Encontrar la(s) otra(s) parte(s)
                    other_parts = [p for p in parts if p != fixed_pair_str]
                    # ASUMPCIÓN: Solo nos interesa comparar a través de UNA otra VI
                    if len(other_parts) == 1:
                        try:
                            other_part = other_parts[0]
                            other_vi_name, other_desc = other_part.split('=')
                            other_alias = aliases.get(other_desc, other_desc)
                            # El display name es solo la parte variable
                            display_name = f"{other_vi_name}: {other_alias}"
                            # Guardar la clave original asociada a este display
                            filtered_groups[key] = display_name
                        except ValueError:
                             logger.warning(f"Error parseando parte variable '{other_parts}' de la clave '{key}'")
                    elif len(other_parts) > 1:
                         logger.warning(f"Modo 2VIs: Se encontraron múltiples partes variables para clave '{key}' con fijo '{fixed_pair_str}'. Se omite por ahora.")
                    # else: No hay otra parte? Caso raro, ignorar.

        logger.debug(f"Grupos filtrados resultantes: {filtered_groups}")
        # Devolver ordenado por display name para consistencia en UI
        return dict(sorted(filtered_groups.items(), key=lambda item: item[1]))


    def get_common_columns_for_groups(self, study_id: int, frequency: str, calculation: str, group_keys: list[str]) -> list[str]:
        """
        Encuentra las columnas de datos comunes presentes en las tablas de resumen
        discreto para una combinación específica de frecuencia, cálculo y grupos.

        :param study_id: ID del estudio.
        :param frequency: Tipo de Dato (ej: 'Cinematica').
        :param calculation: Cálculo (ej: 'Maximo').
        :param group_keys: Lista de claves de grupo (ej: ['CMJ_PRE', 'CMJ_POST']).
        :return: Lista de nombres de columnas comunes ('Attr/Col/Unit').
                 Retorna lista vacía si no hay comunes o falta archivo.
        """
        logger.debug(f"Buscando columnas comunes para estudio {study_id}, freq={frequency}, calc={calculation}, grupos_config={group_keys}")
        common_columns_set = None
        tables_path = self.get_discrete_analysis_tables_path(study_id)

        if not tables_path or not group_keys:
            return []

        freq_path = tables_path / frequency
        if not freq_path.exists():
            logger.warning(f"Directorio de frecuencia no encontrado: {freq_path}")
            return []

        # Determinar si estamos en modo 1VI (main effect) o 2VIs/combinado
        # Heurística: si alguna clave NO contiene ';', es probable que sea modo 1VI.
        # Y si es modo 1VI, group_keys contendrá claves parciales como "VI=Sub-valor".
        is_main_effect_mode_heuristic = any(';' not in key for key in group_keys)
        
        keys_of_tables_to_inspect = set()

        if is_main_effect_mode_heuristic:
            logger.debug("Detectado modo de efecto principal (heurística) para encontrar columnas comunes.")
            # group_keys son parciales, ej: ["Edad=joven", "Edad=mayor"]
            all_full_combined_keys_map, _ = self._identify_study_groups(study_id, frequency) # dict: file_base -> full_key
            all_full_combined_keys_set = set(all_full_combined_keys_map.values())

            for partial_key_selected in group_keys: # ej: "Edad=joven"
                for full_key in all_full_combined_keys_set:
                    if partial_key_selected in full_key.split(';'):
                        keys_of_tables_to_inspect.add(full_key)
            if not keys_of_tables_to_inspect:
                logger.warning(f"Modo efecto principal: No se encontraron tablas de resumen completas para las claves parciales: {group_keys}")
                return []
        else:
            logger.debug("Detectado modo 2VIs/combinado para encontrar columnas comunes.")
            # group_keys son claves completas
            keys_of_tables_to_inspect.update(group_keys)

        logger.debug(f"Se inspeccionarán las siguientes claves de tabla (completas): {keys_of_tables_to_inspect}")

        for full_group_key in keys_of_tables_to_inspect:
            safe_group_key_part = full_group_key.replace('=', '_').replace(';', '__')
            table_filename = f"{calculation}_{frequency}_{safe_group_key_part}.csv"
            table_path = freq_path / table_filename

            if not table_path.exists():
                logger.warning(f"Tabla de resumen no encontrada al buscar columnas comunes: {table_path} (esperada para clave completa '{full_group_key}'). No se pueden determinar columnas comunes.")
                return [] # Si falta CUALQUIER tabla relevante, no hay columnas comunes garantizadas.

            try:
                df_header = pd.read_csv(table_path, sep=',', decimal='.', encoding='utf-8', header=[0, 1, 2], index_col=0, nrows=0)
                current_table_columns = set()
                for header_tuple in df_header.columns:
                    try:
                        if len(header_tuple) == 3:
                            attr, col, unit = map(str, header_tuple)
                            current_table_columns.add(f"{attr}/{col}/{unit}")
                        else:
                            logger.warning(f"Tupla de cabecera con longitud inesperada en {table_filename}: {header_tuple}. Se omitirá.")
                    except Exception as e_header:
                        logger.warning(f"Error procesando tupla de cabecera {header_tuple} en {table_filename}: {e_header}. Se omitirá.")
                
                logger.debug(f"Columnas parseadas de {table_filename}: {current_table_columns}")

                if common_columns_set is None:
                    common_columns_set = current_table_columns
                else:
                    common_columns_set.intersection_update(current_table_columns)

                if not common_columns_set: # Si la intersección se vuelve vacía
                    logger.warning(f"No se encontraron columnas comunes después de procesar {table_filename}")
                    return []
            except Exception as e:
                logger.error(f"Error leyendo cabeceras de {table_path}: {e}", exc_info=True)
                return []

        if common_columns_set is None:
            return []
        return sorted(list(common_columns_set))

    def perform_individual_analysis(self, study_id: int, config: dict):
        """
        Realiza un análisis individual basado en la configuración, genera un gráfico
        y guarda la configuración.

        :param study_id: ID del estudio.
        :param config: Diccionario con la configuración del análisis:
                       {'name': str, 'frequency': str, 'calculation': str,
                        'column': str, 'groups': list[str],
                        'parametric': bool, 'paired': bool}
        :return: Diccionario con rutas al gráfico y archivo de config generados.
                 {'plot_path': str, 'config_path': str}
        :raises ValueError: Si la configuración es inválida o faltan datos/archivos.
        :raises Exception: Si ocurre un error durante el análisis o graficación.
        """
        analysis_name_log = config.get('name', 'N/A')
        logger.info(f"Inicio perform_individual_analysis para estudio {study_id}: {analysis_name_log}") # LOG INICIO
        logger.debug(f"Configuración recibida: {config}")

        # --- Validación de Configuración ---
        required_keys = ['name', 'frequency', 'calculation', 'column',
                         'groups', 'parametric', 'paired']
        if not all(key in config for key in required_keys):
            raise ValueError("Configuración de análisis incompleta.")
        if len(config['groups']) < 2:
            raise ValueError("Se requieren al menos dos grupos para comparar.")
        if not config['name'] or not config['name'].strip():
            raise ValueError("El nombre del análisis no puede estar vacío.")
        # Validar caracteres inválidos en el nombre
        analysis_name = config['name'].strip()
        invalid_chars = r'<>:"/\|?*'
        if any(char in analysis_name for char in invalid_chars):
            raise ValueError(f"El nombre del análisis contiene caracteres "
                             f"inválidos: {invalid_chars}")

        # --- Preparar Rutas ---
        study_path = self.file_service._get_study_path(study_id)
        if not study_path:
            raise ValueError(f"No se pudo encontrar ruta del estudio {study_id}.")

        # --- Nueva estructura de carpetas para Análisis Discreto Individual ---
        variable_analizada_full_discrete = config.get('column', 'VariableDesconocida')
        parts_discrete = variable_analizada_full_discrete.split('/')
        variable_folder_name_discrete_parts = parts_discrete[:2]
        variable_folder_name_discrete = " ".join(variable_folder_name_discrete_parts).replace("/", "_")

        analysis_output_dir = (study_path / "Analisis Discreto" / "Graficos" / 
                               variable_folder_name_discrete / analysis_name)
        try:
            analysis_output_dir.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            raise OSError(f"No se pudo crear directorio de salida para análisis:"
                          f" {analysis_output_dir}\n{e}")

        plot_path = analysis_output_dir / "boxplot.png"
        config_path = analysis_output_dir / "config.json"

        # --- Leer Datos ---
        frequency = config['frequency']
        calculation = config['calculation']
        column_str = config['column']
        mode = config.get("grouping_mode")
        primary_vi_name_for_main_effect = config.get("primary_vi_name") if mode == "1VI" else None
        # groups_from_config son las claves seleccionadas en el diálogo.
        # Para modo 1VI, estas son AHORA parciales (ej: ["Edad=Joven", "Edad=Mayor"])
        # Para modo 2VIs o modo combinado, estas son completas (ej: ["Edad=Joven;Peso=NP", "Edad=Mayor;Peso=OS"])
        groups_from_config = config['groups']


        # --- Validación y Parseo de Columna (maneja '/' en unidad) ---
        try:
            target_column_parts = column_str.split('/', 2)
            if len(target_column_parts) != 3:
                raise ValueError("Formato incorrecto, faltan separadores '/'")
            target_multi_index_col = tuple(target_column_parts)
            logger.debug(f"Columna parseada para análisis: {target_multi_index_col}")
        except Exception as e:
            raise ValueError(f"Formato de columna inválido: '{column_str}'. Error: {e}")

        tables_path = self.get_discrete_analysis_tables_path(study_id)
        if not tables_path:
            raise FileNotFoundError(f"Directorio de tablas de análisis discreto no encontrado para estudio {study_id}")
        freq_path = tables_path / frequency
        if not freq_path.exists():
            raise FileNotFoundError(f"Directorio de frecuencia '{frequency}' no encontrado en {tables_path}")

        data_by_group = [] # Esta será la lista final de datos para cada grupo a comparar
        actual_group_keys_for_legend = [] # Claves/nombres que se usarán para la leyenda

        logger.debug(f"Inicio lectura de datos. Modo: {mode}, Primary VI (1VI): {primary_vi_name_for_main_effect}")

        if mode == "1VI" and primary_vi_name_for_main_effect:
            logger.info(f"Procesando en modo 1VI (efecto principal) para VI: '{primary_vi_name_for_main_effect}'")
            # groups_from_config son AHORA las claves PARCIALES seleccionadas por el usuario,
            # ej: ["Edad=Joven", "Edad=Mayor"]
            
            all_full_keys_in_study_map, _ = self._identify_study_groups(study_id, frequency)
            all_available_full_keys_in_study_set = set(all_full_keys_in_study_map.values())

            for selected_partial_key_from_config in groups_from_config: # e.g., "Edad=Joven"
                pooled_data_for_this_group = []
                descriptor_for_legend = selected_partial_key_from_config.split('=',1)[1] # Get "Joven" from "Edad=Joven"
                actual_group_keys_for_legend.append(descriptor_for_legend)

                unique_full_keys_to_pool_from = set()
                for full_key_in_study in all_available_full_keys_in_study_set: 
                    if selected_partial_key_from_config in full_key_in_study.split(';'): 
                        unique_full_keys_to_pool_from.add(full_key_in_study)
                
                if not unique_full_keys_to_pool_from:
                    logger.warning(f"No se encontraron tablas de resumen completas que contengan la clave parcial '{selected_partial_key_from_config}' en modo 1VI.")
                
                for full_key_to_load in unique_full_keys_to_pool_from: # e.g., "Edad=Joven;Peso=NP", then "Edad=Joven;Peso=OS"
                    safe_full_key_part = full_key_to_load.replace('=', '_').replace(';', '__')
                    table_filename = f"{calculation}_{frequency}_{safe_full_key_part}.csv"
                    table_path = freq_path / table_filename
                    logger.debug(f"Modo 1VI: Leyendo tabla '{table_path}' para agrupar bajo clave parcial '{selected_partial_key_from_config}' (tabla original: '{full_key_to_load}')")

                    if table_path.exists():
                        try:
                            df = pd.read_csv(table_path, sep=',', decimal='.', encoding='utf-8', header=[0, 1, 2], index_col=0)
                            if target_multi_index_col not in df.columns:
                                logger.warning(f"Columna '{column_str}' no encontrada en tabla {table_filename} (para clave parcial '{selected_partial_key_from_config}'). Se omite esta tabla para la agrupación.")
                                continue
                            
                            data_from_table = df[target_multi_index_col].dropna().tolist()
                            pooled_data_for_this_group.extend(data_from_table)
                        except Exception as e:
                            logger.error(f"Error procesando tabla {table_path} (para clave parcial '{selected_partial_key_from_config}'): {e}", exc_info=True)
                    else:
                        logger.warning(f"Modo 1VI: Tabla de resumen no encontrada: {table_path} (esperada para clave parcial '{selected_partial_key_from_config}', tabla original '{full_key_to_load}')")
                
                data_by_group.append(pooled_data_for_this_group)
                logger.info(f"Modo 1VI: Para grupo de efecto principal '{descriptor_for_legend}' (basado en VI '{primary_vi_name_for_main_effect}'), se recolectaron {len(pooled_data_for_this_group)} puntos de datos de {len(unique_full_keys_to_pool_from)} tablas.")

        else: # Modo 2VIs o modo combinado (claves en groups_from_config son completas)
            logger.info(f"Procesando en modo 2VIs o combinado. Claves de grupo: {groups_from_config}")
            actual_group_keys_for_legend = groups_from_config # Las claves completas se usarán para generar leyendas después
            for group_key in groups_from_config: # group_key es una clave completa, ej: "VI1=DescA;VI2=DescB"
                safe_group_key_part = group_key.replace('=', '_').replace(';', '__')
                table_filename = f"{calculation}_{frequency}_{safe_group_key_part}.csv"
                table_path = freq_path / table_filename
                logger.debug(f"Modo 2VIs/Combinado: Intentando leer tabla para grupo '{group_key}': {table_path}")

                if not table_path.exists():
                    logger.error(f"No se encontró tabla resumen requerida: {table_path}")
                    # Intentar listar archivos en el directorio para depuración
                    try:
                        existing_files = [f.name for f in freq_path.iterdir() if f.is_file()]
                        logger.debug(f"Archivos existentes en {freq_path}: {existing_files}")
                    except Exception as list_e:
                        logger.debug(f"No se pudo listar archivos en {freq_path}: {list_e}")
                    raise FileNotFoundError(f"No se encontró tabla resumen requerida: {table_path}")

                try:
                    df = pd.read_csv(table_path, sep=',', decimal='.', encoding='utf-8', header=[0, 1, 2], index_col=0)
                    if target_multi_index_col not in df.columns:
                        raise ValueError(f"Columna '{column_str}' no encontrada en tabla {table_filename}")
                    
                    group_data = df[target_multi_index_col].dropna().tolist()
                    if not group_data:
                        logger.warning(f"No se encontraron datos válidos para grupo '{group_key}' y columna '{column_str}' en {table_filename}")
                    data_by_group.append(group_data)
                except Exception as e:
                    logger.error(f"Error procesando tabla {table_path} para grupo '{group_key}': {e}", exc_info=True)
                    raise ValueError(f"Error leyendo datos para grupo {group_key}: {e}")

        if not any(data_by_group):
            raise ValueError("No se encontraron datos válidos en ninguna tabla para los grupos y columna seleccionados.")

        # --- Generar Título y Leyendas Específicos del Modo ---
        aliases = self.study_service.get_study_aliases(study_id)
        # mode, primary_vi_name_for_main_effect, fixed_vi, fixed_desc_display ya están definidos arriba
        fixed_vi = config.get("fixed_vi_name") # Para modo 2VIs
        fixed_desc_display = config.get("fixed_descriptor_display") # Para modo 2VIs

        title = f"{calculation.capitalize()} de {column_str}"
        group_legend_names = []
        group_xaxis_labels = []

        # Reconstruir leyendas basadas en actual_group_keys_for_legend y el modo
        if mode == "1VI" and primary_vi_name_for_main_effect:
            title += f" (Comparando niveles de {primary_vi_name_for_main_effect})"
            for i, desc_for_legend in enumerate(actual_group_keys_for_legend): # desc_for_legend es ej: "joven"
                alias = aliases.get(desc_for_legend, desc_for_legend)
                group_legend_names.append(f"{primary_vi_name_for_main_effect}: {alias}")
                group_xaxis_labels.append(f"G{i+1}")
        elif mode == "2VIs" and fixed_vi and fixed_desc_display:
            title += f" ({fixed_vi}: {fixed_desc_display})"
            # actual_group_keys_for_legend son las claves completas, ej: "VI2=DescX;VI_Fija=ValorFijo"
            # Necesitamos extraer la parte variable para la leyenda
            fixed_desc_original = fixed_desc_display.split(" (")[0] # Obtener valor original del sub-valor fijo
            fixed_pair_str_to_remove = f"{fixed_vi}={fixed_desc_original}"

            for i, full_key in enumerate(actual_group_keys_for_legend):
                variable_part_display = []
                for part in full_key.split(';'):
                    if part != fixed_pair_str_to_remove:
                        try:
                            vi_name, desc_val = part.split('=',1)
                            alias = aliases.get(desc_val, desc_val)
                            variable_part_display.append(f"{vi_name}: {alias}")
                        except ValueError:
                            variable_part_display.append(part) # Fallback
                legend_name = ", ".join(variable_part_display) if variable_part_display else f"Grupo {i+1}"
                group_legend_names.append(legend_name)
                group_xaxis_labels.append(f"G{i+1}")
        else: # Modo combinado o fallback
            title += " (Comparación de Grupos Combinados)"
            # actual_group_keys_for_legend son las claves completas
            all_display_groups_tuples = self.get_discrete_analysis_groups(study_id, frequency)
            all_display_groups_map = {key: display for display, key in all_display_groups_tuples}
            for i, full_key in enumerate(actual_group_keys_for_legend):
                full_display_name = all_display_groups_map.get(full_key, full_key)
                # Extraer solo el nombre base del display name completo ("Grupo X - Nombre Base")
                base_display = full_display_name.split(" - ", 1)[1] if " - " in full_display_name else full_display_name
                group_legend_names.append(base_display)
                group_xaxis_labels.append(f"G{i+1}")


        # Asegurar que data_by_group esté ordenado según group_legend_names si es necesario
        # En este punto, data_by_group y group_legend_names/group_xaxis_labels se construyen en el mismo orden.

        # Obtener ylabel del gráfico (unidad)
        try:
            chart_ylabel = f"{calculation.capitalize()} ({target_column_parts[2]})" # target_column_parts definido antes
        except IndexError:
            chart_ylabel = f"{calculation.capitalize()}"


        logger.info(f"Título del gráfico: {title}")
        logger.info(f"Etiquetas Eje X: {group_xaxis_labels}")
        logger.info(f"Leyendas del gráfico: {group_legend_names}")


        # --- Realizar Análisis Estadístico ---
        stats_results = None
        if stats:  # Verificar si scipy.stats está disponible
                try:
                    n_groups = len(data_by_group)
                    is_paired = config['paired']
                    is_parametric = config['parametric']
                    test_name = "N/A"
                    p_value = np.nan

                    if n_groups == 2:
                        group1_data = np.array(data_by_group[0])
                        group2_data = np.array(data_by_group[1])

                        if is_paired:
                            # Asegurar misma longitud para tests pareados
                            min_len = min(len(group1_data), len(group2_data))
                            if min_len < 1:
                                raise ValueError("Datos insuficientes para "
                                                 "test pareado.")
                            group1_data = group1_data[:min_len]
                            group2_data = group2_data[:min_len]

                            if is_parametric:
                                test_name = "T-test relacionado"
                                _, p_value = stats.ttest_rel(
                                    group1_data, group2_data, nan_policy='omit'
                                )
                            else:
                                test_name = "Wilcoxon signed-rank"
                                # Wilcoxon requiere > 0 diferencias
                                try:
                                    _, p_value = stats.wilcoxon(
                                        group1_data, group2_data,
                                        nan_policy='omit'
                                    )
                                except ValueError as e:
                                    logger.warning(f"No se pudo ejecutar Wilcoxon "
                                                   f"para {analysis_name_log}: {e}")
                                    p_value = np.nan  # Marcar como no calculable
                        else:  # Independiente
                            if is_parametric:
                                test_name = "T-test independiente"
                                # Log data before test
                                logger.debug(f"Datos para {test_name} (Grupo 1, n={len(group1_data)}): {group1_data[:10]}...") # Muestra primeros 10
                                logger.debug(f"Datos para {test_name} (Grupo 2, n={len(group2_data)}): {group2_data[:10]}...") # Muestra primeros 10
                                # Welch's t-test por defecto
                                try:
                                    stat_result, p_value = stats.ttest_ind(
                                        group1_data, group2_data, equal_var=False,
                                        nan_policy='omit'
                                    )
                                    logger.debug(f"{test_name} resultado: stat={stat_result}, p={p_value}")
                                except Exception as ttest_e:
                                    logger.warning(f"Error ejecutando {test_name} para {analysis_name_log}: {ttest_e}")
                                    p_value = np.nan # Marcar como no calculable si falla
                            else:
                                test_name = "Mann-Whitney U"
                                # Log data before test
                                logger.debug(f"Datos para {test_name} (Grupo 1, n={len(group1_data)}): {group1_data[:10]}...")
                                logger.debug(f"Datos para {test_name} (Grupo 2, n={len(group2_data)}): {group2_data[:10]}...")
                                _, p_value = stats.mannwhitneyu(
                                    group1_data, group2_data,
                                    alternative='two-sided', nan_policy='omit'
                                )

                    elif n_groups > 2:
                        # Filtrar grupos vacíos antes de pasar a ANOVA/Kruskal
                        valid_data_for_test = [np.array(g) for g in data_by_group
                                               if len(g) > 0]
                        if len(valid_data_for_test) < 2:
                            raise ValueError("Se necesitan al menos dos grupos "
                                             "con datos para comparar.")

                        if is_paired:  # ANOVA medidas repetidas / Friedman
                            # Nota: ANOVA medidas repetidas es más complejo.
                            # Usar Friedman como alternativa no paramétrica.
                            if is_parametric:
                                test_name = "ANOVA MR (NO IMPL.)"
                                logger.warning(f"{test_name} para {analysis_name_log}. "
                                               f"Usando Friedman como fallback.")
                                try:
                                    _, p_value = stats.friedmanchisquare(
                                        *valid_data_for_test
                                    )
                                    test_name = "Friedman (fallback)"
                                except ValueError as e:
                                    logger.warning(f"No se pudo ejecutar Friedman "
                                                   f"para {analysis_name_log}: {e}")
                                    p_value = np.nan
                            else:
                                test_name = "Friedman"
                                try:
                                    _, p_value = stats.friedmanchisquare(
                                        *valid_data_for_test
                                    )
                                except ValueError as e:
                                    logger.warning(f"No se pudo ejecutar Friedman "
                                                   f"para {analysis_name_log}: {e}")
                                    p_value = np.nan
                        else:  # ANOVA de un factor / Kruskal-Wallis
                            if is_parametric:
                                test_name = "ANOVA (un factor)"
                                _, p_value = stats.f_oneway(*valid_data_for_test)
                            else:
                                test_name = "Kruskal-Wallis"
                                _, p_value = stats.kruskal(*valid_data_for_test,
                                                           nan_policy='omit')

                    if not np.isnan(p_value):
                        stats_results = {'test_name': test_name,
                                         'p_value': p_value}
                        logger.info(f"Análisis estadístico para {analysis_name_log}: "
                                    f"{test_name}, p-valor = {p_value:.4f}")
                    else:
                        logger.warning(f"No se pudo calcular p-valor para "
                                       f"{analysis_name_log} con test {test_name}.")

                except ValueError as ve:
                    logger.error(f"Error en datos para análisis estadístico de "
                                 f"{analysis_name_log}: {ve}")
                except Exception as e_stat:
                    logger.error(f"Error inesperado durante análisis estadístico "
                                 f"de {analysis_name_log}: {e_stat}",
                                 exc_info=True)
        else:
            logger.warning("Scipy no encontrado. Omitiendo pruebas estadísticas.")

        # --- Generar Gráfico ---
        try:
            # Usar título y leyendas generados arriba
            # Pasar los datos ordenados
            charting.create_comparison_boxplot(
                data_by_group=data_by_group, # Usar datos ordenados
                group_xaxis_labels=group_xaxis_labels, # Etiquetas cortas G1, G2...
                group_legend_names=group_legend_names, # Leyendas específicas del modo
                title=title, # Título específico del modo
                ylabel=chart_ylabel,
                output_path=plot_path,
                stats_results=stats_results
            )
            logger.info(f"Gráfico boxplot generado en: {plot_path}")
        except Exception as e:
            logger.error(f"Error generando gráfico para análisis "
                         f"{analysis_name}: {e}", exc_info=True)
            raise Exception(f"Error generando el gráfico: {e}")

        # --- Guardar Gráfico Interactivo (HTML) ---
        interactive_plot_path = analysis_output_dir / "boxplot_interactive.html"
        try:
            # Pasar datos ordenados y etiquetas/leyendas correctas
            charting.create_interactive_comparison_boxplot(
                data_by_group=data_by_group, # Usar datos ordenados
                group_xaxis_labels=group_xaxis_labels, # Etiquetas cortas G1, G2...
                group_legend_names=group_legend_names, # Leyendas específicas del modo
                title=title, # Título específico del modo
                ylabel=chart_ylabel,
                output_path=interactive_plot_path
            )
        except Exception as e_plotly:
            logger.error(f"Error generando gráfico interactivo para análisis "
                         f"{analysis_name}: {e_plotly}", exc_info=True)
            # No relanzar, el estático ya se generó, pero marcar como no disponible
            interactive_plot_path = None


        # --- Guardar Configuración (incluyendo resultados estadísticos) ---
        config_to_save = config.copy()
        if stats_results:
            # Guardar solo test y p-valor para evitar problemas de serialización
            config_to_save['stats_results'] = {
                'test_name': stats_results.get('test_name', 'N/A'),
                'p_value': stats_results.get('p_value') # Puede ser NaN
            }
        else:
             config_to_save['stats_results'] = None # Indicar que no hubo test

        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                # Usar json.dump con manejo de NaN
                json.dump(config_to_save, f, indent=4, allow_nan=True)
            # Pequeña pausa para asegurar visibilidad del archivo
            time.sleep(0.1)
            logger.info(f"Configuración análisis guardada en: {config_path}")
        except Exception as e:
            logger.error(f"Error guardando config análisis {analysis_name}: {e}", exc_info=True)
            # No relanzar necesariamente, el gráfico ya se generó
            # Considerar eliminar gráfico si falla guardado de config?

        result_paths = {
            'plot_path': str(plot_path),
            'config_path': str(config_path)
        }
        if interactive_plot_path and interactive_plot_path.exists():
            result_paths['interactive_plot_path'] = str(interactive_plot_path)

        return result_paths

    def _get_individual_analysis_base_dir(self, study_id: int) -> Path | None:
        """Obtiene el directorio base para los análisis individuales."""
        study_path = self.file_service._get_study_path(study_id)
        if not study_path:
            logger.error(f"No se pudo encontrar ruta estudio {study_id} para "
                         f"análisis individual.")
            return None
        # La base ahora es "Graficos", pero la estructura completa depende de la variable.
        # Este método podría necesitar la variable si se quiere la ruta completa.
        # Por ahora, devolvemos la base de "Graficos".
        return study_path / "Analisis Discreto" / "Graficos"

    def does_individual_analysis_exist(self, study_id: int, variable_folder_name: str, analysis_name: str) -> bool:
        """
        Verifica si ya existe un análisis discreto individual con el nombre dado para una variable específica.

        :param study_id: ID del estudio.
        :param variable_folder_name: Nombre de la carpeta de la variable (ej: "LAnkleAngles X").
        :param analysis_name: Nombre del análisis a verificar.
        :return: True si existe, False en caso contrario.
        """
        base_graphics_dir = self._get_individual_analysis_base_dir(study_id)
        if not base_graphics_dir or not base_graphics_dir.exists():
            return False
        
        # Sanitize analysis_name just in case, though it should be clean from dialog
        clean_analysis_name = analysis_name.strip()
        invalid_chars = r'<>:"/\|?*' # Mismos caracteres inválidos que en continuo
        if any(char in clean_analysis_name for char in invalid_chars):
            logger.warning(f"Intento de verificar existencia de análisis individual con nombre inválido: '{clean_analysis_name}'")
            return False

        # Path: .../Analisis Discreto/Graficos/[VARIABLE_FOLDER_NAME]/[ANALYSIS_NAME]
        analysis_path = base_graphics_dir / variable_folder_name / clean_analysis_name
        return analysis_path.exists() and analysis_path.is_dir()

    # get_individual_analysis_path is removed. The UI will pass the full path.

    def list_individual_analyses(self, study_id: int) -> list[dict]:
        """
        Lista los análisis individuales guardados para un estudio.

        :param study_id: ID del estudio.
        :return: Lista de diccionarios:
                 {'name': str, 'path': Path, 'config': dict, 'mtime': float,
                  'plot_path': Path, 'interactive_plot_path': Path | None}
        """
        analyses = []
        # La base ahora es .../Analisis Discreto/Graficos/
        base_graphics_dir = self._get_individual_analysis_base_dir(study_id)
        if not base_graphics_dir or not base_graphics_dir.exists():
            return []

        # Iterar sobre las subcarpetas de variables (ej. "LAnkleMoment X")
        for variable_folder_path in base_graphics_dir.iterdir():
            if not variable_folder_path.is_dir():
                continue
            
            # Ahora iterar sobre las carpetas de análisis dentro de cada variable_folder
            for analysis_specific_folder_path in variable_folder_path.iterdir():
                if analysis_specific_folder_path.is_dir():
                    analysis_name = analysis_specific_folder_path.name
                    config_path = analysis_specific_folder_path / "config.json"
                    plot_path = analysis_specific_folder_path / "boxplot.png"
                    interactive_plot_path = analysis_specific_folder_path / "boxplot_interactive.html"

                    if config_path.exists() and config_path.is_file():
                        try:
                            with open(config_path, 'r', encoding='utf-8') as f:
                                config_data = json.load(f)
                            mtime = config_path.stat().st_mtime
                            analyses.append({
                                'name': analysis_name,
                                'path': analysis_specific_folder_path, # Ruta a la carpeta del análisis específico
                                'config': config_data,
                                'mtime': mtime,
                                'plot_path': plot_path,
                                'interactive_plot_path': interactive_plot_path if interactive_plot_path.exists() else None
                            })
                        except json.JSONDecodeError:
                            logger.error(f"Error leyendo config.json para análisis '{analysis_name}' en {analysis_specific_folder_path}.")
                        except Exception as e:
                            logger.error(f"Error procesando análisis '{analysis_name}' en {analysis_specific_folder_path}: {e}", exc_info=True)
                    elif not config_path.exists() or not config_path.is_file():
                        logger.warning(f"Directorio análisis '{analysis_name}' en {analysis_specific_folder_path} encontrado sin config.json válido.")

        # Ordenar por fecha de modificación (más reciente primero)
        analyses.sort(key=lambda x: x['mtime'], reverse=True)
        return analyses

    def _delete_individual_analysis_no_backup(self, analysis_path_to_delete: Path):
        """
        Core logic to delete an individual analysis folder. No backup trigger here.
        """
        # Backup is handled by public method or batch operation
        logger.info(f"Core logic: eliminar análisis individual en: {analysis_path_to_delete}")

        if not analysis_path_to_delete.exists():
            raise FileNotFoundError(f"El directorio del análisis no existe: {analysis_path_to_delete}")
        if not analysis_path_to_delete.is_dir():
            raise ValueError(f"La ruta del análisis no es un directorio: {analysis_path_to_delete}")

        try:
            shutil.rmtree(analysis_path_to_delete)
            logger.info(f"Análisis individual eliminado: {analysis_path_to_delete}")

            # Limpiar directorio padre (variable_folder_name) si queda vacío.
            parent_dir = analysis_path_to_delete.parent
            if parent_dir.exists() and parent_dir.is_dir() and not any(parent_dir.iterdir()):
                # Verificar que el nombre del directorio padre no sea "Graficos" directamente
                # para evitar eliminar la carpeta base "Graficos" si es el último análisis en general.
                if parent_dir.name != "Graficos":
                    logger.info(f"Directorio de variable '{parent_dir.name}' está vacío tras eliminar '{analysis_path_to_delete.name}', eliminándolo.")
                    try:
                        parent_dir.rmdir() # rmdir solo funciona si está vacío
                        logger.info(f"Directorio de variable '{parent_dir.name}' eliminado.")
                    except OSError as e_rmparent:
                        logger.error(f"Error eliminando directorio padre vacío '{parent_dir.name}': {e_rmparent}", exc_info=True)
                else:
                    logger.debug(f"Directorio padre es '{parent_dir.name}' (carpeta base 'Graficos'), no se eliminará aunque esté vacío.")
            elif parent_dir.exists() and parent_dir.is_dir(): # Parent exists but is not empty
                 logger.debug(f"Directorio de variable '{parent_dir.name}' no está vacío, no se eliminará.")
            # else: parent_dir does not exist or is not a dir, nothing to do.

        except OSError as e:
            logger.error(f"Error eliminando directorio análisis {analysis_path_to_delete}: {e}", exc_info=True)
            raise

    def delete_individual_analysis(self, analysis_path_to_delete: Path):
        """
        Elimina la carpeta y contenido de un análisis individual específico, triggering backup.
        """
        try:
            create_backup(backup_type='automatic')
        except Exception as e_backup:
            logger.error(f"Error creating automatic backup before deleting individual analysis {analysis_path_to_delete.name}: {e_backup}", exc_info=True)
            # Log and continue for now
        
        if self.undo_manager.is_undo_enabled():
            if not self.undo_manager.prepare_undo_cache():
                logger.error(f"Failed to prepare undo cache for deleting individual analysis {analysis_path_to_delete}. Aborting delete operation.")
                raise Exception(f"Failed to prepare undo cache for deleting individual analysis {analysis_path_to_delete}. Deletion aborted.")

            if not self.undo_manager.cache_item_for_undo(str(analysis_path_to_delete), "individual_analysis_folder"):
                logger.warning(f"Failed to cache individual analysis folder {analysis_path_to_delete} for undo. Deletion will proceed but undo might be partial.")
        
        try:
            self._delete_individual_analysis_no_backup(analysis_path_to_delete)
        except Exception as e:
            logger.error(f"Error during deletion of individual analysis {analysis_path_to_delete} after undo preparation: {e}", exc_info=True)
            raise

    def delete_all_discrete_summary_tables(self, study_id: int):
        """
        Elimina todas las tablas de resumen discreto (.xlsx) para un estudio.
        :param study_id: ID del estudio.
        """
        try:
            create_backup(backup_type='automatic')
        except Exception as e_backup:
            logger.error(f"Error creating automatic backup before deleting all discrete summary tables for study {study_id}: {e_backup}", exc_info=True)
            # Log and continue for now

        tables_base_dir = self.get_discrete_analysis_tables_path(study_id)
        if not tables_base_dir or not tables_base_dir.exists():
            logger.info(f"No se encontró directorio de tablas de resumen para estudio {study_id}. Nada que eliminar.")
            return 0 # No files deleted

        # Collect all .xlsx and their corresponding .csv files to be deleted for caching
        items_to_cache_for_undo = [] # List of tuples (path_str, item_type)
        if self.undo_manager.is_undo_enabled():
            for freq_dir_cache in tables_base_dir.iterdir():
                if freq_dir_cache.is_dir():
                    for table_file_cache in freq_dir_cache.iterdir():
                        if table_file_cache.is_file() and table_file_cache.suffix == '.xlsx':
                            items_to_cache_for_undo.append((str(table_file_cache), "discrete_summary_table_xlsx"))
                            # Check for corresponding .csv file
                            csv_file_cache = table_file_cache.with_suffix('.csv')
                            if csv_file_cache.exists() and csv_file_cache.is_file():
                                items_to_cache_for_undo.append((str(csv_file_cache), "discrete_summary_table_csv"))
            
        if self.undo_manager.is_undo_enabled():
            # Always prepare cache if undo is enabled for this operation.
            # This call clears the previous cache and backs up the database.
            if not self.undo_manager.prepare_undo_cache():
                logger.error(f"Failed to prepare undo cache for deleting all discrete tables in study {study_id}. Aborting delete operation.")
                raise Exception(f"Failed to prepare undo cache for deleting all discrete tables in study {study_id}. Deletion aborted.")
            
            # If cache preparation was successful, proceed to cache identified items
            if items_to_cache_for_undo:
                for path_str, item_type in items_to_cache_for_undo:
                    if not self.undo_manager.cache_item_for_undo(path_str, item_type):
                        logger.warning(f"Failed to cache discrete table item {path_str} (type: {item_type}) for undo during 'delete all'. Deletion will proceed but undo might be partial.")
            else:
                # This case means undo is enabled, cache is prepared (DB backed up, old cache cleared),
                # but no specific files from *this* operation needed caching.
                logger.info(f"Undo cache prepared for 'delete all discrete tables', but no .xlsx or .csv tables found to cache in {tables_base_dir}.")
        # If undo is not enabled, no caching actions are taken here; prepare_undo_cache would have returned False.


        deleted_count = 0
        logger.info(f"Iniciando eliminación de todas las tablas de resumen en: {tables_base_dir}")
        # Iterate again for actual deletion
        for freq_dir_delete in list(tables_base_dir.iterdir()): 
            if freq_dir_delete.is_dir():
                for table_file_delete in list(freq_dir_delete.iterdir()):
                    if table_file_delete.is_file() and table_file_delete.suffix == '.xlsx':
                        try:
                            self._delete_discrete_summary_table_no_backup(str(table_file_delete))
                            deleted_count += 1
                        except OSError as e:
                            logger.error(f"Error eliminando tabla de resumen {table_file_delete} (durante delete_all): {e}", exc_info=True)
                # Limpiar carpeta de frecuencia si queda vacía
                if not any(freq_dir_delete.iterdir()):
                    try:
                        freq_dir_delete.rmdir()
                        logger.info(f"Carpeta de frecuencia de tablas vacía eliminada: {freq_dir_delete}")
                    except OSError as e:
                        logger.error(f"Error eliminando carpeta de frecuencia de tablas vacía {freq_dir_delete}: {e}", exc_info=True)
        
        # Limpiar carpeta base "Tablas" si queda vacía
        if tables_base_dir.exists() and not any(tables_base_dir.iterdir()):
            try:
                tables_base_dir.rmdir()
                logger.info(f"Carpeta base de tablas de resumen vacía eliminada: {tables_base_dir}")
            except OSError as e:
                 logger.error(f"Error eliminando carpeta base de tablas vacía {tables_base_dir}: {e}", exc_info=True)
        
        logger.info(f"Eliminación de tablas de resumen completada. Total eliminadas: {deleted_count}.")
        return deleted_count


    def delete_all_individual_analyses(self, study_id: int):
        """
        Elimina todos los análisis discretos individuales guardados para un estudio.
        :param study_id: ID del estudio.
        """
        try:
            create_backup(backup_type='automatic')
        except Exception as e_backup:
            logger.error(f"Error creating automatic backup before deleting all individual analyses for study {study_id}: {e_backup}", exc_info=True)
            # Decide if operation should continue or be aborted. For now, logging and continuing.

        analyses_base_dir = self._get_individual_analysis_base_dir(study_id) # .../Analisis Discreto/Graficos
        if not analyses_base_dir or not analyses_base_dir.exists():
            logger.info(f"No se encontró directorio base de análisis individuales para estudio {study_id}. Nada que eliminar.")
            return 0

        folders_to_cache_for_undo = []
        if self.undo_manager.is_undo_enabled():
            for var_folder_cache in analyses_base_dir.iterdir():
                if var_folder_cache.is_dir():
                    for analysis_folder_cache in var_folder_cache.iterdir():
                        if analysis_folder_cache.is_dir():
                            folders_to_cache_for_undo.append(str(analysis_folder_cache))
            
            if folders_to_cache_for_undo:
                if not self.undo_manager.prepare_undo_cache():
                    logger.error(f"Failed to prepare undo cache for deleting all individual analyses in study {study_id}. Aborting delete operation.")
                    raise Exception(f"Failed to prepare undo cache for deleting all individual analyses in study {study_id}. Deletion aborted.")

                for folder_to_cache_str in folders_to_cache_for_undo:
                    if not self.undo_manager.cache_item_for_undo(folder_to_cache_str, "individual_analysis_folder"):
                        logger.warning(f"Failed to cache individual analysis folder {folder_to_cache_str} for undo during 'delete all'. Deletion will proceed but undo might be partial.")
            elif not self.undo_manager.is_undo_enabled():
                pass
            else:
                logger.info(f"Undo enabled, but no individual analysis folders found to cache in {analyses_base_dir} for 'delete all'.")

        deleted_count = 0
        logger.info(f"Iniciando eliminación de todos los análisis individuales en: {analyses_base_dir}")
        for variable_folder_delete in list(analyses_base_dir.iterdir()):
            if variable_folder_delete.is_dir():
                for analysis_folder_delete in list(variable_folder_delete.iterdir()):
                    if analysis_folder_delete.is_dir(): 
                        try:
                            # For "delete all", we call shutil.rmtree directly after caching.
                            # The _delete_individual_analysis_no_backup also calls rmtree,
                            # but this avoids repeated checks if we already iterated for caching.
                            shutil.rmtree(analysis_folder_delete)
                            deleted_count += 1
                            logger.info(f"Análisis individual eliminado: {analysis_folder_delete}")
                        except OSError as e:
                            logger.error(f"Error eliminando análisis individual {analysis_folder_delete}: {e}", exc_info=True)
                # Limpiar carpeta de variable si queda vacía
                if not any(variable_folder_delete.iterdir()):
                    try:
                        variable_folder_delete.rmdir()
                        logger.info(f"Carpeta de variable de análisis individual vacía eliminada: {variable_folder_delete}")
                    except OSError as e:
                        logger.error(f"Error eliminando carpeta de variable vacía {variable_folder_delete}: {e}", exc_info=True)
        
        # Limpiar carpeta base "Graficos" si queda vacía
        if analyses_base_dir.exists() and not any(analyses_base_dir.iterdir()):
            try:
                analyses_base_dir.rmdir()
                logger.info(f"Carpeta base de gráficos de análisis individual vacía eliminada: {analyses_base_dir}")
            except OSError as e:
                logger.error(f"Error eliminando carpeta base de gráficos vacía {analyses_base_dir}: {e}", exc_info=True)

        logger.info(f"Eliminación de análisis individuales completada. Total eliminados: {deleted_count}.")
        return deleted_count

    def delete_selected_individual_analyses(self, analysis_paths: list[Path]) -> tuple[int, list[str]]:
        """
        Elimina una lista de análisis individuales específicos.

        :param analysis_paths: Lista de objetos Path de las carpetas de análisis a eliminar.
        :return: Tupla (número de eliminaciones exitosas, lista de mensajes de error).
        """
        if not analysis_paths:
            return 0, []

        success_count = 0
        errors = []
        try:
            create_backup(backup_type='automatic')
        except Exception as e_backup:
            logger.error(f"Error creating automatic backup before deleting selected individual analyses: {e_backup}", exc_info=True)
            # Log and continue for now

        if self.undo_manager.is_undo_enabled() and analysis_paths:
            if not self.undo_manager.prepare_undo_cache():
                logger.error(f"Failed to prepare undo cache for deleting selected individual analyses. Aborting delete operation.")
                raise Exception(f"Failed to prepare undo cache for deleting selected individual analyses. Deletion aborted.")
            
            for path_to_cache in analysis_paths:
                if not self.undo_manager.cache_item_for_undo(str(path_to_cache), "individual_analysis_folder"):
                    logger.warning(f"Failed to cache individual analysis folder {path_to_cache} for undo during 'delete selected'. Deletion will proceed but undo might be partial.")
        
        for analysis_path in analysis_paths:
            try:
                self._delete_individual_analysis_no_backup(analysis_path) 
                success_count += 1
                logger.info(f"Análisis individual {analysis_path.name} eliminado como parte de una operación masiva.")
            except Exception as e:
                error_msg = f"Error eliminando análisis individual {analysis_path.name}: {e}"
                logger.error(error_msg, exc_info=True)
                errors.append(error_msg)
        
        if errors:
            logger.warning(f"Se encontraron errores durante la eliminación masiva de análisis individuales: {errors}")
        return success_count, errors

    def delete_all_continuous_analyses(self, study_id: int):
        """
        Elimina todos los análisis continuos guardados para un estudio.
        :param study_id: ID del estudio.
        """
        try:
            create_backup(backup_type='automatic')
        except Exception as e_backup:
            logger.error(f"Error creating automatic backup before deleting all continuous analyses for study {study_id}: {e_backup}", exc_info=True)
            # Decide if operation should continue or be aborted. For now, logging and continuing.

        analyses_base_dir = self._get_continuous_analysis_base_dir(study_id) # .../Analisis Continuo
        if not analyses_base_dir or not analyses_base_dir.exists():
            logger.info(f"No se encontró directorio base de análisis continuos para estudio {study_id}. Nada que eliminar.")
            return 0

        folders_to_cache_for_undo_cont = []
        if self.undo_manager.is_undo_enabled():
            for var_folder_cache_cont in analyses_base_dir.iterdir():
                if var_folder_cache_cont.is_dir():
                    for analysis_folder_cache_cont in var_folder_cache_cont.iterdir():
                        if analysis_folder_cache_cont.is_dir():
                            folders_to_cache_for_undo_cont.append(str(analysis_folder_cache_cont))
            
            if folders_to_cache_for_undo_cont:
                if not self.undo_manager.prepare_undo_cache():
                    logger.error(f"Failed to prepare undo cache for deleting all continuous analyses in study {study_id}. Aborting delete operation.")
                    raise Exception(f"Failed to prepare undo cache for deleting all continuous analyses in study {study_id}. Deletion aborted.")

                for folder_to_cache_str_cont in folders_to_cache_for_undo_cont:
                    if not self.undo_manager.cache_item_for_undo(folder_to_cache_str_cont, "continuous_analysis_folder"):
                        logger.warning(f"Failed to cache continuous analysis folder {folder_to_cache_str_cont} for undo during 'delete all'. Deletion will proceed but undo might be partial.")
            elif not self.undo_manager.is_undo_enabled():
                pass
            else:
                logger.info(f"Undo enabled, but no continuous analysis folders found to cache in {analyses_base_dir} for 'delete all'.")

        deleted_count = 0
        logger.info(f"Iniciando eliminación de todos los análisis continuos en: {analyses_base_dir}")
        for variable_folder_delete_cont in list(analyses_base_dir.iterdir()):
            if variable_folder_delete_cont.is_dir():
                for analysis_folder_delete_cont in list(variable_folder_delete_cont.iterdir()):
                    if analysis_folder_delete_cont.is_dir(): 
                        try:
                            shutil.rmtree(analysis_folder_delete_cont)
                            deleted_count += 1
                            logger.info(f"Análisis continuo eliminado: {analysis_folder_delete_cont}")
                        except OSError as e:
                            logger.error(f"Error eliminando análisis continuo {analysis_folder_delete_cont}: {e}", exc_info=True)
                # Limpiar carpeta de variable si queda vacía
                if not any(variable_folder_delete_cont.iterdir()):
                    try:
                        variable_folder_delete_cont.rmdir()
                        logger.info(f"Carpeta de variable de análisis continuo vacía eliminada: {variable_folder_delete_cont}")
                    except OSError as e:
                        logger.error(f"Error eliminando carpeta de variable vacía {variable_folder_delete_cont}: {e}", exc_info=True)

        # Limpiar carpeta base "Analisis Continuo" si queda vacía
        if analyses_base_dir.exists() and not any(analyses_base_dir.iterdir()):
            try:
                analyses_base_dir.rmdir()
                logger.info(f"Carpeta base de análisis continuo vacía eliminada: {analyses_base_dir}")
            except OSError as e:
                logger.error(f"Error eliminando carpeta base de análisis continuo vacía {analyses_base_dir}: {e}", exc_info=True)

        logger.info(f"Eliminación de análisis continuos completada. Total eliminados: {deleted_count}.")
        return deleted_count

    def delete_selected_continuous_analyses(self, analysis_paths: list[Path]) -> tuple[int, list[str]]:
        """
        Elimina una lista de análisis continuos específicos.

        :param analysis_paths: Lista de objetos Path de las carpetas de análisis a eliminar.
        :return: Tupla (número de eliminaciones exitosas, lista de mensajes de error).
        """
        if not analysis_paths:
            return 0, []

        success_count = 0
        errors = []
        try:
            create_backup(backup_type='automatic')
        except Exception as e_backup:
            logger.error(f"Error creating automatic backup before deleting selected continuous analyses: {e_backup}", exc_info=True)
            # Log and continue for now

        if self.undo_manager.is_undo_enabled() and analysis_paths:
            if not self.undo_manager.prepare_undo_cache():
                logger.error(f"Failed to prepare undo cache for deleting selected continuous analyses. Aborting delete operation.")
                raise Exception(f"Failed to prepare undo cache for deleting selected continuous analyses. Deletion aborted.")
            
            for path_to_cache_cont in analysis_paths:
                if not self.undo_manager.cache_item_for_undo(str(path_to_cache_cont), "continuous_analysis_folder"):
                    logger.warning(f"Failed to cache continuous analysis folder {path_to_cache_cont} for undo during 'delete selected'. Deletion will proceed but undo might be partial.")

        for analysis_path in analysis_paths:
            try:
                self._delete_continuous_analysis_no_backup(analysis_path) 
                success_count += 1
                logger.info(f"Análisis continuo {analysis_path.name} eliminado como parte de una operación masiva.")
            except Exception as e:
                error_msg = f"Error eliminando análisis continuo {analysis_path.name}: {e}"
                logger.error(error_msg, exc_info=True)
                errors.append(error_msg)

        if errors:
            logger.warning(f"Se encontraron errores durante la eliminación masiva de análisis continuos: {errors}")
        return success_count, errors

    def delete_selected_discrete_summary_tables(self, table_paths: list[str]) -> tuple[int, list[str]]:
        """
        Elimina una lista de tablas de resumen discretas específicas.

        :param table_paths: Lista de strings de rutas a las tablas .xlsx a eliminar.
        :return: Tupla (número de eliminaciones exitosas, lista de mensajes de error).
        """
        if not table_paths:
            return 0, []

        success_count = 0
        errors = []
        try:
            create_backup(backup_type='automatic')
        except Exception as e_backup:
            logger.error(f"Error creating automatic backup before deleting selected discrete summary tables: {e_backup}", exc_info=True)
            # Log and continue for now

        if self.undo_manager.is_undo_enabled() and table_paths:
            if not self.undo_manager.prepare_undo_cache():
                logger.error(f"Failed to prepare undo cache for deleting selected discrete tables. Aborting delete operation.")
                raise Exception(f"Failed to prepare undo cache for deleting selected discrete tables. Deletion aborted.")
            
            for table_str_to_cache in table_paths: # table_str_to_cache is the path to an .xlsx file
                xlsx_path_to_cache = Path(table_str_to_cache)
                if not self.undo_manager.cache_item_for_undo(str(xlsx_path_to_cache), "discrete_summary_table_xlsx"):
                    logger.warning(f"Failed to cache discrete table {xlsx_path_to_cache} for undo during 'delete selected'. Deletion will proceed but undo might be partial.")
                
                # Also cache the corresponding .csv file if it exists
                csv_path_to_cache = xlsx_path_to_cache.with_suffix('.csv')
                if csv_path_to_cache.exists() and csv_path_to_cache.is_file():
                    if not self.undo_manager.cache_item_for_undo(str(csv_path_to_cache), "discrete_summary_table_csv"):
                        logger.warning(f"Failed to cache corresponding CSV table {csv_path_to_cache} for undo during 'delete selected'. Deletion will proceed but undo might be partial.")

        for table_path_str in table_paths:
            try:
                self._delete_discrete_summary_table_no_backup(table_path_str) 
                success_count += 1
                logger.info(f"Tabla de resumen {Path(table_path_str).name} eliminada como parte de una operación masiva.")
            except Exception as e:
                error_msg = f"Error eliminando tabla de resumen {Path(table_path_str).name}: {e}"
                logger.error(error_msg, exc_info=True)
                errors.append(error_msg)
        
        if errors:
            logger.warning(f"Se encontraron errores durante la eliminación masiva de tablas de resumen: {errors}")
        return success_count, errors
